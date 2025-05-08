"""第三方设备点表配置服务"""
import logging
from typing import List, Dict, Any

# DAO 和领域模型
from .database.dao import ConfiguredDeviceDAO
from .models.configured_device_models import ConfiguredDevicePointModel
from .models.template_models import DeviceTemplateModel # 需要模板模型来配置点

logger = logging.getLogger(__name__)

class ConfigService:
    """处理第三方设备点表配置的生成、持久化和导出。"""

    def __init__(self, config_dao: ConfiguredDeviceDAO):
        if not config_dao:
            logger.error("ConfigService 初始化失败: 未提供 ConfiguredDeviceDAO 实例。")
            raise ValueError("ConfiguredDeviceDAO 实例是必需的")
        self.config_dao = config_dao
        logger.info("ConfigService 初始化完成。")

    def configure_points_from_template(self, template: DeviceTemplateModel, device_prefix: str) -> bool:
        """根据模板和前缀生成设备点位，并保存到数据库。"""
        if not template or not template.points:
            logger.warning("尝试从未包含点位的模板生成配置。")
            return False
        
        configured_points_to_save: List[ConfiguredDevicePointModel] = []
        for point_model in template.points:
            # 创建 ConfiguredDevicePointModel 实例
            configured_point = ConfiguredDevicePointModel(
                template_name=template.name, # 存储模板名称快照
                device_prefix=device_prefix,
                var_suffix=point_model.var_suffix,
                desc_suffix=point_model.desc_suffix,
                data_type=point_model.data_type
                # created_at 会由数据库自动添加, id 也是
            )
            configured_points_to_save.append(configured_point)
            
        try:
            # 调用DAO批量保存
            success = self.config_dao.save_configured_points(configured_points_to_save)
            if success:
                logger.info(f"为模板 '{template.name}' 和前缀 '{device_prefix}' 成功配置并保存了 {len(configured_points_to_save)} 个点位。")
            return success
        except Exception as e:
            # DAO层应该已经记录了错误
            logger.error(f"服务层保存配置点位时发生错误: {e}", exc_info=True)
            return False

    def get_all_configured_points(self) -> List[ConfiguredDevicePointModel]:
        """从数据库获取所有已配置的设备点位。"""
        try:
            return self.config_dao.get_all_configured_points()
        except Exception as e:
            logger.error(f"服务层获取所有已配置点位失败: {e}", exc_info=True)
            return []

    def clear_all_configurations(self) -> bool:
        """从数据库删除所有已配置的设备点位。"""
        try:
            return self.config_dao.delete_all_configured_points()
        except Exception as e:
            logger.error(f"服务层清空所有配置失败: {e}", exc_info=True)
            return False

    def get_configuration_summary(self) -> List[Dict]:
        """获取配置摘要信息 (用于UI显示)。"""
        try:
            # 从DAO获取原始分组数据
            raw_summary = self.config_dao.get_configuration_summary_raw()
            
            # 格式化为UI需要的结构
            summary_list = []
            for row in raw_summary:
                summary_list.append({
                    'template': row['template_name'],
                    'variable': row['device_prefix'], # UI中使用'variable'作为键
                    'count': row['point_count'],
                    'status': "已配置" # 状态总是"已配置"
                })
            return summary_list
        except Exception as e:
            logger.error(f"服务层获取配置摘要失败: {e}", exc_info=True)
            return []

    def export_to_excel(self, file_path: str) -> None:
        """将所有已配置的点位导出到Excel文件。"""
        try:
            from openpyxl import Workbook # 延迟导入
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("导出Excel需要安装 openpyxl 库: pip install openpyxl")
            raise ImportError("导出功能需要 openpyxl 库。请先安装它。")
            
        points_to_export = self.get_all_configured_points()
        
        if not points_to_export:
            raise ValueError("没有可导出的点位数据")
            
        # 新的分组方式：按模板名称分组
        grouped_by_template_name: Dict[str, List[ConfiguredDevicePointModel]] = {}
        for point in points_to_export:
            key = point.template_name # 只使用模板名作为键
            grouped_by_template_name.setdefault(key, []).append(point)
            
        wb = Workbook()
        wb.remove(wb.active) # 移除默认创建的空sheet
        
        sheet_count = 0
        # 遍历按模板名称分组的数据
        for template_name, points_list in grouped_by_template_name.items():
            # 生成Sheet标题，主要基于模板名称
            safe_tmpl_name = "".join(c for c in template_name if c.isalnum() or c in (' ', '_', '-'))
            sheet_title_base = safe_tmpl_name 
            sheet_title = sheet_title_base[:31] # Excel限制31字符
            
            # 简单的重名处理，以防万一（例如模板名处理后恰好相同或特殊字符被移除导致重名）
            original_sheet_title = sheet_title
            count = 1
            while sheet_title in wb.sheetnames: # wb.sheetnames 是当前工作簿中所有sheet的名称列表
                suffix = f"_{count}"
                max_len = 31 - len(suffix)
                sheet_title = original_sheet_title[:max_len] + suffix
                count += 1
                if count > 100: # 防止死循环
                     raise ValueError(f"无法为模板'{template_name}'创建唯一Sheet名称")

            ws = wb.create_sheet(title=sheet_title)
            
            # 写入表头
            headers = ["变量名", "描述", "数据类型"]
            ws.append(headers)
            
            # 写入数据行 (points_list 中包含该模板名下的所有点，无论设备前缀)
            for point in points_list:
                ws.append([
                    point.variable_name, # 使用计算属性 (已包含前缀)
                    point.description,   # 使用计算属性 (已包含前缀)
                    point.data_type
                ])
                
            # 自动调整列宽
            for i, column_header in enumerate(headers):
                column_letter = get_column_letter(i + 1)
                max_length = len(column_header) # Start with header length
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except:
                        pass # Ignore errors
                adjusted_width = (max_length + 2) * 1.1 # Add padding and factor
                ws.column_dimensions[column_letter].width = adjusted_width
            
            sheet_count += 1
            
        if sheet_count == 0:
             raise ValueError("分组后没有可导出的点位数据") # 再次检查

        try:
            wb.save(file_path)
            logger.info(f"点表已成功导出至 {file_path}")
        except Exception as save_e:
            logger.error(f"保存Excel文件失败: {file_path}, 错误: {save_e}", exc_info=True)
            raise IOError(f"保存Excel文件失败: {save_e}") from save_e 