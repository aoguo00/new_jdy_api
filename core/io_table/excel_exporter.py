import logging
from typing import List, Dict, Any, Optional, Callable
import re # 新增导入

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

# 尝试导入 openpyxl 及其组件，以便在类定义中引用类型提示
try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    # 如果导入失败，定义一个占位符，以便类型提示不会引发错误
    # 实际的错误处理将在 IOExcelExporter 的 export_to_excel 方法中进行
    Worksheet = Any 
    Font = Any
    Alignment = Any
    PatternFill = Any
    Border = Any
    Side = Any
    get_column_letter = Any
    openpyxl = None


class PLCAddressAllocator:
    """负责PLC地址的分配和管理。"""
    def __init__(self, start_md_address=320, start_mx_byte=20, start_mx_bit=0):
        self.current_md_address = start_md_address
        self.current_mx_byte = start_mx_byte
        self.current_mx_bit = start_mx_bit
        logger.info(f"PLCAddressAllocator initialized: MD starts at {self.current_md_address}, MX starts at {self.current_mx_byte}.{self.current_mx_bit}")

    def allocate_real_address(self) -> str:
        """分配一个REAL类型的地址 (%MD)。"""
        address = f"%MD{self.current_md_address}"
        self.current_md_address += 4
        # logger.debug(f"Allocated REAL address: {address}")
        return address

    def allocate_bool_address(self) -> str:
        """分配一个BOOL类型的地址 (%MX)。"""
        address = f"%MX{self.current_mx_byte}.{self.current_mx_bit}"
        # logger.debug(f"Allocating BOOL address: {address} (before increment: byte={self.current_mx_byte}, bit={self.current_mx_bit})")
        self.current_mx_bit += 1
        if self.current_mx_bit > 7:
            self.current_mx_bit = 0
            self.current_mx_byte += 1
            # logger.debug(f"BOOL address bit overflowed. New byte: {self.current_mx_byte}, new bit: {self.current_mx_bit}")
        return address


class BaseSheetExporter:
    """
    工作表导出器的基类，提供通用的列宽自适应方法。
    """
    # 预定义的列宽配置
    COLUMN_WIDTH_CONFIG = {
        # PLC IO表的列宽配置
        "序号": 5,
        "模块名称": 10,
        "模块类型": 10,
        "供电类型（有源/无源）": 25,
        "线制": 8,
        "通道位号": 15,
        "场站名": 20,
        "场站编号": 20,
        "变量名称（HMI）": 25,
        "变量描述": 35,
        "数据类型": 12,
        "读写属性": 10,
        "保存历史": 10,
        "掉电保护": 10,
        "量程低限": 12,
        "量程高限": 12,
        "SLL设定值": 12,
        "SLL设定点位": 15,
        "SLL设定点位_PLC地址": 25,
        "SLL设定点位_通讯地址": 25,
        "SL设定值": 12,
        "SL设定点位": 15,
        "SL设定点位_PLC地址": 25,
        "SL设定点位_通讯地址": 25,
        "SH设定值": 12,
        "SH设定点位": 15,
        "SH设定点位_PLC地址": 25,
        "SH设定点位_通讯地址": 25,
        "SHH设定值": 12,
        "SHH设定点位": 15,
        "SHH设定点位_PLC地址": 25,
        "SHH设定点位_通讯地址": 25,
        "LL报警": 10,
        "LL报警_PLC地址": 20,
        "LL报警_通讯地址": 20,
        "L报警": 10,
        "L报警_PLC地址": 20,
        "L报警_通讯地址": 20,
        "H报警": 10,
        "H报警_PLC地址": 20,
        "H报警_通讯地址": 20,
        "HH报警": 10,
        "HH报警_PLC地址": 20,
        "HH报警_通讯地址": 20,
        "维护值设定": 12,
        "维护值设定点位": 20,
        "维护值设定点位_PLC地址": 25,
        "维护值设定点位_通讯地址": 25,
        "维护使能开关点位": 20,
        "维护使能开关点位_PLC地址": 25,
        "维护使能开关点位_通讯地址": 26,
        "PLC绝对地址": 20,
        "上位机通讯地址": 20,
        
        # 第三方设备表的列宽配置
        "MODBUS地址": 15,
    }

    def _adjust_column_widths(self, ws: Worksheet, headers: List[str]):
        """
        根据预定义的配置设置列宽。
        """
        if not openpyxl:
            return

        for col_idx, header in enumerate(headers):
            column_letter = get_column_letter(col_idx + 1)
            # 使用预定义的列宽，如果没有预定义则使用默认值
            width = self.COLUMN_WIDTH_CONFIG.get(header, 15)  # 默认宽度为15
            ws.column_dimensions[column_letter].width = width


class PLCSheetExporter(BaseSheetExporter):
    """
    负责生成 "IO点表" (PLC IO数据) Sheet页的核心逻辑。

    该类通过协调一系列私有辅助方法和配置规则（MODULE_PROCESSING_RULES）来处理每个PLC IO点数据，
    最终在提供的 openpyxl Worksheet 对象中填充格式化的IO点表。

    主要职责包括：
    - 初始化并写入表头。
    - 逐行处理传入的 `plc_io_data`。
    - 对每个IO点：
        - 初始化基础行数据。
        - 根据模块类型（如 AI, AO, DI, DO）和 MODULE_PROCESSING_RULES 配置，动态分配所需的PLC地址（绝对地址及特定用途地址）。
        - 根据配置，为特定模块（当前主要是AI）的某些列自动生成基于HMI变量名的Excel公式。
        - 根据配置，将分配的PLC地址及其计算得到的Modbus通讯地址填充到正确的列中。
        - 应用预定义的行样式（如高亮用户输入列、添加边框等）。
    - 调整最终表格的列宽以适应内容。

    `MODULE_PROCESSING_RULES` 是一个关键的类属性配置字典，它定义了不同模块类型
    在PLC地址分配、Excel公式生成和地址列映射方面的特定行为，使得类具有良好的可扩展性。
    """
    # Columns that always require user input (需要用户输入的通用列的表头名称)
    ALWAYS_HIGHLIGHT_HEADERS = {
        "供电类型（有源/无源）",
        "线制",
        "变量名称（HMI）",
        "变量描述"
    }

    # Columns that require user input specifically for AI modules (AI模块特有的需要用户输入的列的表头名称)
    AI_SPECIFIC_HIGHLIGHT_HEADERS = {
        "量程低限",
        "量程高限",
        "SLL设定值",
        "SL设定值",
        "SH设定值",
        "SHH设定值"
    }

    # MODULE_PROCESSING_RULES 配置字典详解：
    # -------------------------------------
    # 该字典定义了不同PLC模块类型（如 "AI", "AO", "DI", "DO"）在生成IO点表时
    # 所遵循的特定处理规则。还包含一个 "_COMMON_" 键，用于定义所有模块通用的规则。
    # 未来若要支持新的模块类型或修改现有模块的行为，主要修改此配置即可。
    #
    # 结构：
    # { "<模块类型字符串>": { # 例如 "AI", "AO", "_COMMON_"
    #       "plc_allocations": [ # (可选) 定义此模块类型除了绝对地址外，还需要分配哪些特定用途的PLC地址。
    #                            # 每个条目是一个元组: (逻辑名称:str, 分配器方法名:str, 结果字典键名:str)
    #                            # - 逻辑名称: 仅为可读性，当前未使用。
    #                            # - 分配器方法名: PLCAddressAllocator 类中用于分配此类地址的方法名 (如 "allocate_real_address")。
    #                            # - 结果字典键名: 分配到的地址在 _allocate_addresses 返回的字典中所使用的键。
    #                          ],
    #       "excel_formulas": [  # (可选) 定义哪些列需要基于 "变量名称（HMI）" 自动生成Excel公式。
    #                           # 每个条目是一个元组: (目标列头名:str, HMI名称后缀:str)
    #                           # - 目标列头名: self.headers_plc 中定义的列名。
    #                           # - HMI名称后缀: 拼接到HMI变量名后的字符串，用于构成完整的点位名。
    #                         ],
    #       "address_mapping": [ # (可选) 定义已分配的PLC地址如何映射到Excel的PLC地址列和通讯地址列。
    #                          # 每个条目是一个元组: (结果字典键名:str, PLC地址列头名:str, 通讯地址列头名:str)
    #                          # - 结果字典键名: _allocate_addresses 返回的字典中的键。
    #                          # - PLC地址列头名: self.headers_plc 中定义的PLC地址列的列名。
    #                          # - 通讯地址列头名: self.headers_plc 中定义的对应通讯地址列的列名。
    #                        ]
    #     }, ...
    # }
    MODULE_PROCESSING_RULES = {
        "AI": {
            "plc_allocations": [
                # 结构: (逻辑名称, 分配器方法名, 结果字典键名)
                ("sll_set", "allocate_real_address", "sll_set_plc_addr"),
                ("sl_set", "allocate_real_address", "sl_set_plc_addr"),
                ("sh_set", "allocate_real_address", "sh_set_plc_addr"),
                ("shh_set", "allocate_real_address", "shh_set_plc_addr"),
                ("ll_alarm", "allocate_bool_address", "ll_alarm_plc_addr"),
                ("l_alarm", "allocate_bool_address", "l_alarm_plc_addr"),
                ("h_alarm", "allocate_bool_address", "h_alarm_plc_addr"),
                ("hh_alarm", "allocate_bool_address", "hh_alarm_plc_addr"),
                ("maint_val_set", "allocate_real_address", "maint_val_set_plc_addr"),
                ("maint_enable", "allocate_bool_address", "maint_enable_plc_addr"),
            ],
            "excel_formulas": [
                # 结构: (目标列头名, HMI名称后缀)
                ("SLL设定点位", "_LoLoLimit"),
                ("SL设定点位", "_LoLimit"),
                ("SH设定点位", "_HiLimit"),
                ("SHH设定点位", "_HiHiLimit"),
                ("LL报警", "_LL"),
                ("L报警", "_L"),
                ("H报警", "_H"),
                ("HH报警", "_HH"),
                ("维护值设定点位", "_whz"),
                ("维护使能开关点位", "_whzzt"),
            ],
            "address_mapping": [ 
                ('sll_set_plc_addr', "SLL设定点位_PLC地址", "SLL设定点位_通讯地址"),
                ('sl_set_plc_addr', "SL设定点位_PLC地址", "SL设定点位_通讯地址"),
                ('sh_set_plc_addr', "SH设定点位_PLC地址", "SH设定点位_通讯地址"),
                ('shh_set_plc_addr', "SHH设定点位_PLC地址", "SHH设定点位_通讯地址"),
                ('ll_alarm_plc_addr', "LL报警_PLC地址", "LL报警_通讯地址"),
                ('l_alarm_plc_addr', "L报警_PLC地址", "L报警_通讯地址"),
                ('h_alarm_plc_addr', "H报警_PLC地址", "H报警_通讯地址"),
                ('hh_alarm_plc_addr', "HH报警_PLC地址", "HH报警_通讯地址"),
                ('maint_val_set_plc_addr', "维护值设定点位_PLC地址", "维护值设定点位_通讯地址"),
                ('maint_enable_plc_addr', "维护使能开关点位_PLC地址", "维护使能开关点位_通讯地址"),
            ]
        },
        "AO": {
            "plc_allocations": [
                # ("maint_val_set", "allocate_real_address", "maint_val_set_plc_addr"),
                # ("maint_enable", "allocate_bool_address", "maint_enable_plc_addr"),
            ],
            "excel_formulas": [], 
            "address_mapping": [
                # ('maint_val_set_plc_addr', "维护值设定点位_PLC地址", "维护值设定点位_通讯地址"),
                # ('maint_enable_plc_addr', "维护使能开关点位_PLC地址", "维护使能开关点位_通讯地址"),
            ]
        },
        "DI": {
            "plc_allocations": [], 
            "excel_formulas": [],
            "address_mapping": []
        },
        "DO": {
            "plc_allocations": [], 
            "excel_formulas": [],
            "address_mapping": []
        },
        "_COMMON_": { 
             "address_mapping": [
                 ('plc_absolute_addr', "PLC绝对地址", "上位机通讯地址")
             ]
        }
    }

    def __init__(self):
        self.headers_plc = [
            "序号", "模块名称", "模块类型", "供电类型（有源/无源）", "线制", "通道位号", # 0-5
            "场站名", "场站编号", "变量名称（HMI）", "变量描述", "数据类型", "读写属性", # 6-11
            "保存历史", "掉电保护", "量程低限", "量程高限", "SLL设定值", "SLL设定点位", # 12-17
            "SLL设定点位_PLC地址", "SLL设定点位_通讯地址", "SL设定值", "SL设定点位", # 18-21
            "SL设定点位_PLC地址", "SL设定点位_通讯地址", "SH设定值", "SH设定点位", # 22-25
            "SH设定点位_PLC地址", "SH设定点位_通讯地址", "SHH设定值", "SHH设定点位", # 26-29
            "SHH设定点位_PLC地址", "SHH设定点位_通讯地址", "LL报警", "LL报警_PLC地址", # 30-33
            "LL报警_通讯地址", "L报警", "L报警_PLC地址", "L报警_通讯地址", "H报警", # 34-38
            "H报警_PLC地址", "H报警_通讯地址", "HH报警", "HH报警_PLC地址", "HH报警_通讯地址", # 39-43
            "维护值设定", "维护值设定点位", "维护值设定点位_PLC地址", "维护值设定点位_通讯地址", # 44-47
            "维护使能开关点位", "维护使能开关点位_PLC地址", "维护使能开关点位_通讯地址", # 48-50
            "PLC绝对地址", "上位机通讯地址" # 51-52
        ] # 共53列

    def _get_modbus_address(self, plc_address: str) -> str:
        """
        根据PLC地址计算Modbus通讯地址。
        规则:
        - %MDx: Modbus地址 = (x // 2) + 3000 + 40000 + 1 
          (其中40000是特定项目的偏移量，3000是MD区映射到Modbus的基地址，+1是Modbus地址从1开始的调整)
        - %MXm.n: Modbus地址 = (m * 8) + n + 3000 + 1 
          (3000是MX区映射到Modbus的基地址，+1是调整)
        如果PLC地址为空或无法识别，则返回空字符串。
        """
        if not plc_address: # 如果PLC地址为空，直接返回空字符串
            return ""

        md_match = re.fullmatch(r"%MD(\d+)", plc_address)
        if md_match:
            try:
                val = int(md_match.group(1))
                # 规则: (x // 2) + 3000 (MD基地址) + 40000 (特定项目偏移) + 1 (Modbus从1开始)
                return str((val // 2) + 3000 + 40000 + 1) 
            except ValueError:
                logger.warning(f"无法解析%MD地址中的数字: {plc_address}")
                return ""

        # logger.debug(f"尝试匹配MX地址: {plc_address}")
        mx_match = re.fullmatch(r"%MX(\d+)\.(\d+)", plc_address) # 注意转义点号
        if mx_match:
            # logger.debug(f"MX地址匹配成功: Groups={mx_match.groups()}")
            try:
                m_val = int(mx_match.group(1)) # 字节部分
                n_val = int(mx_match.group(2)) # 位部分
                # 规则: (字节 * 8 + 位) + 3000 (MX基地址) + 1 (Modbus从1开始)
                comm_addr = (m_val * 8) + n_val + 3000 + 1
                # logger.debug(f"计算得到的MX通讯地址: {comm_addr} for {plc_address}")
                return str(comm_addr)
            except ValueError:
                logger.warning(f"无法解析%MX地址中的数字: {plc_address}, Groups={mx_match.groups()}")
                return ""
        # else:
            # logger.debug(f"MX地址匹配失败: {plc_address}")
        
        # 这条日志现在应该只在两种类型都不匹配时执行
        logger.debug(f"未识别的PLC地址格式 (非MD也非MX)，无法转换为Modbus地址: {plc_address}") 
        return ""

    def _initialize_row_data(self, point_data: Dict[str, Any], idx: int, site_name: Optional[str], site_no: Optional[str]) -> tuple[List[Any], str, str]:
        """
        初始化行数据列表并填充基础信息。
        返回: (final_row_data, channel_io_type, data_type_value)
        """
        final_row_data = ["" for _ in range(len(self.headers_plc))]

        # 0. 序号 (从1开始)
        final_row_data[0] = idx
        # 1. 模块名称 (来自原始数据)
        final_row_data[1] = point_data.get('model', '')
        # 2. 模块类型 (来自原始数据)
        channel_io_type = point_data.get('type', '')
        final_row_data[2] = channel_io_type
        
        # 3. 供电类型（有源/无源）- 用户填写，高亮
        # 4. 线制 - 用户填写，高亮
        
        # 5. 通道位号 (来自原始数据)
        final_row_data[5] = point_data.get('address', '')
        # 6. 场站名 (来自传入参数)
        final_row_data[6] = site_name if site_name else ""
        # 7. 场站编号 (来自传入参数)
        final_row_data[7] = site_no if site_no else ""
        
        # 8. 变量名称（HMI）- 用户填写，高亮，Excel公式会引用此列
        final_row_data[8] = "" 

        # 9. 变量描述 - 用户填写，高亮
        final_row_data[9] = point_data.get('description', '') # 允许预填，但仍标记为用户输入
        
        # 10. 数据类型 (根据模块类型推断)
        data_type_value = ""
        if channel_io_type in ["AI", "AO"]:
            data_type_value = "REAL"
        elif channel_io_type in ["DI", "DO"]:
            data_type_value = "BOOL"
        final_row_data[10] = data_type_value
        
        # 11. 读写属性 (硬编码为 R/W)
        final_row_data[11] = "R/W"
        # 12. 保存历史 (硬编码为 是)
        final_row_data[12] = "是"  
        # 13. 掉电保护 (硬编码为 是)
        final_row_data[13] = "是"  
        
        # 14. 量程低限 - AI模块用户填写，高亮
        # 15. 量程高限 - AI模块用户填写，高亮
        # 16. SLL设定值 - AI模块用户填写，高亮
        # ... 其他设定值和报警名称列将由后续方法根据HMI名称填充Excel公式 ...

        return final_row_data, channel_io_type, data_type_value

    def _allocate_addresses(self, channel_io_type: str, data_type_value: str, address_allocator: PLCAddressAllocator) -> Dict[str, str]:
        """
        根据模块类型和数据类型分配所有相关的PLC地址。
        返回一个字典，键是地址的逻辑名称，值是分配到的PLC地址字符串。
        例如: {
            'plc_absolute_addr': '%MD320',
            'sll_set_plc_addr': '%MD324',
            'll_alarm_plc_addr': '%MX20.0',
            ...
        }
        """
        allocated_addrs: Dict[str, str] = {}

        # 1. 分配PLC绝对地址 (所有模块类型都需要)
        if data_type_value == "REAL":
            allocated_addrs['plc_absolute_addr'] = address_allocator.allocate_real_address()
        elif data_type_value == "BOOL":
            allocated_addrs['plc_absolute_addr'] = address_allocator.allocate_bool_address()
        else:
            allocated_addrs['plc_absolute_addr'] = "" # 未知数据类型则为空

        # 2. 根据模块IO类型分配其他特定用途的PLC地址
        # DI 和 DO 模块除了绝对地址外，目前不主动分配其他特定用途的PLC地址
        # 如果未来DI/DO需要特定地址，可在此处添加 elif channel_io_type == "DI": ...

        # 使用配置驱动的地址分配
        module_rules = self.MODULE_PROCESSING_RULES.get(channel_io_type, {})
        specific_allocations = module_rules.get("plc_allocations", [])

        for _, allocator_method_name, addr_key in specific_allocations:
            if hasattr(address_allocator, allocator_method_name):
                allocator_func = getattr(address_allocator, allocator_method_name)
                allocated_addrs[addr_key] = allocator_func()
            else:
                logger.warning(f"PLCAddressAllocator 中未找到方法: {allocator_method_name} (模块类型: {channel_io_type})")

        return allocated_addrs

    def _populate_module_formulas(self, final_row_data: List[Any], idx: int, channel_io_type: str):
        """
        根据模块类型配置，为相应的点位/报警名称列填充基于HMI名称的Excel公式。
        直接修改传入的 final_row_data。
        idx: 当前数据行的1-based索引，用于构建正确的Excel单元格引用。
        channel_io_type: 当前处理的模块IO类型。
        """
        module_rules = self.MODULE_PROCESSING_RULES.get(channel_io_type, {})
        formula_rules = module_rules.get("excel_formulas", [])

        if not formula_rules: # 如果没有此模块的公式规则，则直接返回
            return

        # "变量名称（HMI）" 在 headers_plc 中的索引是固定的 (8)
        # Excel 列标是 get_column_letter(索引 + 1)
        # Excel 数据行号是 idx (enumerate的起始值) + 1 (因为表头占了第1行)
        # (注意：openpyxl.utils.get_column_letter 是1-indexed)
        hmi_name_column_index = self.headers_plc.index("变量名称（HMI）") 
        hmi_name_column_letter = get_column_letter(hmi_name_column_index + 1)
        # populate_sheet 中 enumerate 从 1 开始计数，所以 idx 直接对应 Excel 数据区的行号
        # 表头是第1行，数据从第2行开始。如果 idx 是 enumerate(..., 1) 的结果，
        # 那么第一条数据行的 Excel 行号是 idx + 1 (因为表头占了第一行)。
        excel_data_row_number = idx + 1 
        hmi_name_cell_ref = f"{hmi_name_column_letter}{excel_data_row_number}"

        for target_column_header, suffix in formula_rules:
            try:
                target_column_idx = self.headers_plc.index(target_column_header)
                final_row_data[target_column_idx] = f'={hmi_name_cell_ref} & "{suffix}"'
            except ValueError:
                logger.warning(f"在表头中未找到列: {target_column_header} (模块类型: {channel_io_type}，用于Excel公式生成)")

    def _fill_addresses_into_row(self, final_row_data: List[Any], allocated_plc_addrs: Dict[str, str], channel_io_type: str):
        """
        将已分配的PLC地址及其对应的Modbus通讯地址填充到final_row_data的相应列中。
        使用配置驱动的地址映射规则。
        直接修改传入的 final_row_data。
        allocated_plc_addrs: 由 _allocate_addresses 方法返回的PLC地址字典。
        channel_io_type: 当前处理的模块IO类型。
        """
        module_rules = self.MODULE_PROCESSING_RULES.get(channel_io_type, {})
        specific_address_mapping = module_rules.get("address_mapping", [])
        
        common_address_mapping = self.MODULE_PROCESSING_RULES.get("_COMMON_", {}).get("address_mapping", [])
        
        # 合并特定模块的映射和通用映射
        # 注意：如果特定模块的映射与通用映射有冲突（不太可能，因为键名不同），特定模块的会覆盖（如果简单合并列表）
        # 这里我们选择分别迭代，或者确保键名不冲突所以顺序不重要
        all_mappings = specific_address_mapping + common_address_mapping

        if not all_mappings:
            return

        for addr_key, plc_col_header, comm_col_header in all_mappings:
            plc_addr = allocated_plc_addrs.get(addr_key, "")
            try:
                plc_col_idx = self.headers_plc.index(plc_col_header)
                comm_col_idx = self.headers_plc.index(comm_col_header)
                
                final_row_data[plc_col_idx] = plc_addr
                final_row_data[comm_col_idx] = self._get_modbus_address(plc_addr)
            except ValueError:
                logger.warning(f"在表头中未找到列: '{plc_col_header}' 或 '{comm_col_header}' (模块类型: {channel_io_type}，用于地址填充)")

    def _apply_row_styles(self, 
                          ws: Worksheet, 
                          current_row_num: int, 
                          channel_io_type: str, 
                          user_input_fill: Optional[PatternFill],
                          thin_border_style: Optional[Border],
                          left_alignment: Optional[Alignment]):
        """
        为指定行中的单元格应用高亮、边框和对齐样式。
        ws: 当前操作的工作表。
        current_row_num: 需要应用样式的行号。
        channel_io_type: 当前行的模块IO类型，用于条件高亮。
        user_input_fill: 用于高亮用户输入单元格的填充样式。
        thin_border_style: 用于单元格的边框样式。
        left_alignment: 用于单元格的左对齐样式。
        """
        for col_idx, header_title in enumerate(self.headers_plc):
            cell = ws.cell(row=current_row_num, column=col_idx + 1)
            
            # 1. 应用高亮 (如果user_input_fill已定义)
            if user_input_fill:
                should_highlight = False
                if header_title in self.ALWAYS_HIGHLIGHT_HEADERS:
                    should_highlight = True
                elif header_title in self.AI_SPECIFIC_HIGHLIGHT_HEADERS and channel_io_type == "AI":
                    should_highlight = True
                
                if should_highlight:
                    cell.fill = user_input_fill
            
            # 2. 应用边框 (如果thin_border_style已定义)
            if thin_border_style:
                cell.border = thin_border_style
        
        # 3. 应用左对齐到整行 (如果left_alignment已定义)
        # 注意: openpyxl 中对整行应用样式通常是遍历该行的所有单元格
        if left_alignment: 
            for cell_in_row in ws[current_row_num]: 
                cell_in_row.alignment = left_alignment

    def populate_sheet(self, ws: Worksheet, plc_io_data: List[Dict[str, Any]], site_name: Optional[str], site_no: Optional[str]):
        """
        核心方法：填充PLC IO数据到指定的工作表。
        协调各个辅助方法完成数据处理、地址分配、名称生成和样式应用。
        """
        if not openpyxl: return # 确保openpyxl已加载

        # --- 初始化样式对象 --- 
        left_alignment = None
        thin_border_style = None
        highlight_fill_color = "FFE4E1E1" # 用户输入高亮颜色 (浅灰色)
        user_input_fill = None

        if Alignment and Alignment is not Any:
            left_alignment = Alignment(horizontal='left', vertical='center')
        if Border and Side and Border is not Any and Side is not Any:
            thin_side = Side(border_style="thin", color="000000") # 细黑边
            thin_border_style = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        if PatternFill and PatternFill is not Any:
            user_input_fill = PatternFill(start_color=highlight_fill_color, end_color=highlight_fill_color, fill_type="solid")

        # --- 1. 写入表头并应用样式 --- 
        ws.append(self.headers_plc)
        for cell in ws[1]: # 表头行
            cell.font = Font(bold=True) # 加粗
            if left_alignment:
                cell.alignment = left_alignment # 左对齐
            if thin_border_style:
                cell.border = thin_border_style # 应用边框

        # --- 2. 初始化PLC地址分配器 --- 
        address_allocator = PLCAddressAllocator()

        # --- 3. 遍历IO点数据，逐行处理和填充 --- 
        for idx, point_data in enumerate(plc_io_data, 1): # idx从1开始，对应Excel中的行号（数据区）
            
            # 3.1 初始化行数据并填充基础信息
            final_row_data, channel_io_type, data_type_value = self._initialize_row_data(point_data, idx, site_name, site_no)

            # 3.2 分配所有相关的PLC地址
            allocated_plc_addresses = self._allocate_addresses(channel_io_type, data_type_value, address_allocator)

            # 3.3 为AI模块填充Excel公式 (如果适用)
            self._populate_module_formulas(final_row_data, idx, channel_io_type)
            
            # 3.4 将分配的PLC地址和计算出的通讯地址填充到行数据中
            self._fill_addresses_into_row(final_row_data, allocated_plc_addresses, channel_io_type)

            # --- 3.X 追加已填充完毕的行数据到工作表 --- 
            ws.append(final_row_data)

            # --- 3.Y 应用高亮和边框样式到刚添加的行 --- 
            current_excel_row = ws.max_row # 获取当前写入的行号 (即追加数据后的最大行号)
            self._apply_row_styles(ws, 
                                 current_excel_row, 
                                 channel_io_type, 
                                 user_input_fill, 
                                 thin_border_style, 
                                 left_alignment)
        
        # --- 4. 调整所有列的宽度 --- 
        self._adjust_column_widths(ws, self.headers_plc)
        return address_allocator # 返回最终的地址分配器实例


class ThirdPartySheetExporter(BaseSheetExporter):
    """
    负责生成第三方设备点表的Sheet页。
    """
    def __init__(self):
        super().__init__() # 调用父类的构造函数，如果BaseSheetExporter有__init__
        self.headers_tp = [
            "场站名", "变量名称", "变量描述", "数据类型", 
            "SLL设定值", "SL设定值", "SH设定值", "SHH设定值",  # 新增和调整顺序后的设定值列
            "PLC地址", "MODBUS地址"
        ]
        # self.get_modbus_address_func = None # 移除旧的成员变量方式

    def populate_sheet(self, ws: Worksheet, points_in_template: List[Dict[str, Any]], site_name: Optional[str], address_allocator: PLCAddressAllocator, get_modbus_address_func: Callable[[str], str]):
        """
        填充第三方设备数据到指定的工作表，并使用传入的地址分配器和转换函数。
        """
        if not openpyxl: return

        left_alignment = None
        if Alignment and Alignment is not Any: 
            left_alignment = Alignment(horizontal='left', vertical='center')

        ws.append(self.headers_tp)
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            if left_alignment:
                cell.alignment = left_alignment

        for tp_point in points_in_template:
            plc_address = ""
            modbus_address = ""
            # points_in_template 中的每个点期望有 'data_type' 字段
            data_type = tp_point.get('data_type', '').upper()

            if data_type == "REAL" or data_type == "INT" or data_type == "DINT": 
                plc_address = address_allocator.allocate_real_address()
            elif data_type == "BOOL":
                plc_address = address_allocator.allocate_bool_address()
            else:
                logger.warning(f"第三方点位 '{tp_point.get('point_name', '未知')}' 数据类型 '{data_type}' 未知或不支持自动分配PLC地址，将留空。")

            if plc_address:
                try:
                    modbus_address = get_modbus_address_func(plc_address)
                except Exception as e:
                    logger.error(f"为PLC地址 '{plc_address}' 计算MODBUS地址时出错: {e}")
                    modbus_address = "计算错误"
            
            row_data_tp = [
                site_name if site_name else "",
                tp_point.get('point_name', ''), # 假设原始数据包含 point_name
                tp_point.get('description', ''),# 假设原始数据包含 description
                tp_point.get('data_type', ''), # 使用原始数据类型以保持大小写一致性（如果需要）
                tp_point.get('sll_setpoint', ""), # SLL设定值
                tp_point.get('sl_setpoint', ""),  # SL设定值
                tp_point.get('sh_setpoint', ""),  # SH设定值
                tp_point.get('shh_setpoint', ""), # SHH设定值
                plc_address,
                modbus_address  
            ]
            ws.append(row_data_tp)
            if left_alignment:
                for cell_in_row in ws[ws.max_row]: 
                    cell_in_row.alignment = left_alignment
        
        self._adjust_column_widths(ws, self.headers_tp)


class IOExcelExporter:
    """
    负责将IO点数据导出到Excel文件。
    使用 PLCSheetExporter 和 ThirdPartySheetExporter 来处理具体的Sheet页生成。
    """

    def __init__(self):
        self.plc_sheet_exporter = PLCSheetExporter()
        self.third_party_sheet_exporter = ThirdPartySheetExporter()

    def export_to_excel(self,
                        plc_io_data: Optional[List[Dict[str, Any]]],
                        third_party_data: Optional[List[Dict[str, Any]]] = None,
                        filename: str = "IO_Table.xlsx",
                        site_name: Optional[str] = None,
                        site_no: Optional[str] = None) -> bool:
        """
        将PLC IO数据和/或第三方设备数据导出到指定的Excel文件。
        """
        logger.info(f"IOExcelExporter.export_to_excel called. filename='{filename}', site_name='{site_name}', site_no='{site_no}'")
        logger.info(f"Received plc_io_data type: {type(plc_io_data)}, length: {len(plc_io_data) if plc_io_data is not None else 'None'}")
        logger.info(f"Received third_party_data type: {type(third_party_data)}, content: {third_party_data}")

        # 模块的导入移到这里，确保在尝试使用前检查
        global openpyxl, Worksheet, Font, get_column_letter, Alignment, PatternFill, Border, Side
        if openpyxl is None: # 检查是否在文件顶部成功导入
            try:
                import openpyxl as opxl_main # 使用别名避免与全局变量冲突
                from openpyxl.worksheet.worksheet import Worksheet as OpxlWorksheet
                from openpyxl.styles import Font as OpxlFont, Alignment as OpxlAlignment, PatternFill as OpxlPatternFill, Border as OpxlBorder, Side as OpxlSide
                from openpyxl.utils import get_column_letter as opxl_get_column_letter
                
                # 更新全局变量
                openpyxl = opxl_main
                Worksheet = OpxlWorksheet
                Font = OpxlFont
                Alignment = OpxlAlignment
                PatternFill = OpxlPatternFill
                Border = OpxlBorder
                Side = OpxlSide
                get_column_letter = opxl_get_column_letter
                logger.info("openpyxl library including PatternFill, Border, Side loaded successfully.")
            except ImportError:
                logger.error("导出Excel失败：openpyxl 库未安装。请运行 'pip install openpyxl'")
                return False

        if not plc_io_data and not third_party_data:
            logger.warning("没有PLC IO数据或第三方设备数据可供导出。")
            return False

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)
            logger.info("Removed default 'Sheet'.")

        # --- 处理PLC IO数据 ---
        shared_address_allocator: Optional[PLCAddressAllocator] = None # 显式声明类型
        if plc_io_data:
            logger.info("Processing PLC IO data...")
            ws_plc = wb.create_sheet(title="IO点表")
            shared_address_allocator = self.plc_sheet_exporter.populate_sheet(ws_plc, plc_io_data, site_name, site_no)
            logger.info("'IO点表' sheet populated.")

        # --- 处理第三方设备数据 ---
        if third_party_data:
            logger.info("Processing third-party data...")
            grouped_tp_data = {}
            for point in third_party_data:
                template_name = point.get('template_name', 'Default_Template') # 提供默认模板名
                if template_name not in grouped_tp_data:
                    grouped_tp_data[template_name] = []
                grouped_tp_data[template_name].append(point)

            for template_name, points_in_template in grouped_tp_data.items():
                safe_sheet_name = template_name.replace('[', '').replace(']', '').replace('*', '').replace('?', '').replace(':', '').replace('/', '').replace('\\\\', '')
                safe_sheet_name = safe_sheet_name[:31] # Excel sheet name length limit
                logger.info(f"尝试为模板 '{template_name}' 创建Sheet，安全名称为: '{safe_sheet_name}'")
                
                ws_tp = wb.create_sheet(title=safe_sheet_name)
                
                # 决定使用哪个地址分配器实例
                allocator_for_tp = shared_address_allocator if shared_address_allocator else PLCAddressAllocator()
                # 传递地址分配器和地址转换函数
                self.third_party_sheet_exporter.populate_sheet(
                    ws_tp, 
                    points_in_template, 
                    site_name, 
                    allocator_for_tp, 
                    self.plc_sheet_exporter._get_modbus_address # 传递方法引用
                )
                logger.info(f"Sheet '{safe_sheet_name}' for template '{template_name}' populated. Current sheets: {wb.sheetnames}")
        
        if not wb.sheetnames:
            logger.error("没有创建任何Sheet页。可能 plc_io_data 和 third_party_data 都为空或处理失败。")
            # openpyxl 在没有可见sheet时保存会出错，确保至少有一个sheet页或明确不保存
            # 但前面的 if not plc_io_data and not third_party_data: 应该已经捕获了这种情况
            # 如果流程到这里还没有sheet，说明逻辑有误，但还是尝试创建一个空sheet避免保存错误
            # 不过更好的做法是如果真的没数据就不尝试保存
            if not plc_io_data and not third_party_data: # 双重检查
                 return False # 明确不保存
            # 如果有数据但sheet创建失败，这是个更深的问题
            # 为防止 'At least one sheet must be visible' 错误，如果到这里还没有sheet，且有数据要写，这是个错误
            # 但我们期望 populate_sheet 总是创建sheet
            # 这里的检查是最后的防线
            logger.warning("Workbook has no sheets, but data was provided. This indicates an issue in sheet creation.")
            #  不应该在这里创建默认sheet，因为这会掩盖问题，如果之前的逻辑正确，这里总会有sheet
            #  如果真的没有sheet，保存会失败，这是期望的行为，以暴露错误。

        try:
            wb.save(filename)
            logger.info(f"数据已成功导出到 {filename}")
            return True
        except Exception as e:
            # 特别处理 openpyxl 因没有可见工作表而引发的错误
            if "At least one sheet must be visible" in str(e):
                logger.error(f"导出Excel文件时出错: {e}. 这通常意味着没有成功创建任何工作表，或者所有工作表都被隐藏了。")
            else:
                logger.error(f"导出Excel文件时出错: {e}")
            return False
