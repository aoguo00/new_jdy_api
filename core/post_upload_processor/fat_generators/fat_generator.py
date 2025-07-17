"""FAT点表生成器模块"""

import os
import logging
from typing import Tuple, Optional
import openpyxl # 导入 openpyxl
from openpyxl.utils import get_column_letter # 用于将列号转为字母（如果需要调试）

logger = logging.getLogger(__name__)

# 定义用户指定的列名
COL_HMI_VARIABLE_NAME = "变量名称（HMI）"
COL_VARIABLE_DESCRIPTION = "变量描述"
COL_CHANNEL_TAG = "通道位号"

# 设定值相关列名
COL_SLL_SET_VALUE = "SLL设定值"
COL_SL_SET_VALUE = "SL设定值"
COL_SH_SET_VALUE = "SH设定值"
COL_SHH_SET_VALUE = "SHH设定值"

# 设定点位及地址列名
COL_SLL_SET_POINT = "SLL设定点位"
COL_SLL_SET_POINT_PLC = "SLL设定点位_PLC地址"
COL_SLL_SET_POINT_COMM = "SLL设定点位_通讯地址"
COL_SL_SET_POINT = "SL设定点位"
COL_SL_SET_POINT_PLC = "SL设定点位_PLC地址"
COL_SL_SET_POINT_COMM = "SL设定点位_通讯地址"
COL_SH_SET_POINT = "SH设定点位"
COL_SH_SET_POINT_PLC = "SH设定点位_PLC地址"
COL_SH_SET_POINT_COMM = "SH设定点位_通讯地址"
COL_SHH_SET_POINT = "SHH设定点位"
COL_SHH_SET_POINT_PLC = "SHH设定点位_PLC地址"
COL_SHH_SET_POINT_COMM = "SHH设定点位_通讯地址"

# 报警列名
COL_LL_ALARM = "LL报警"
COL_LL_ALARM_PLC = "LL报警_PLC地址"
COL_LL_ALARM_COMM = "LL报警_通讯地址"
COL_L_ALARM = "L报警"
COL_L_ALARM_PLC = "L报警_PLC地址"
COL_L_ALARM_COMM = "L报警_通讯地址"
COL_H_ALARM = "H报警"
COL_H_ALARM_PLC = "H报警_PLC地址"
COL_H_ALARM_COMM = "H报警_通讯地址"
COL_HH_ALARM = "HH报警"
COL_HH_ALARM_PLC = "HH报警_PLC地址"
COL_HH_ALARM_COMM = "HH报警_通讯地址"

# 维护相关列名
COL_MAINTENANCE_SET_POINT = "维护值设定点位"
COL_MAINTENANCE_SET_POINT_PLC = "维护值设定点位_PLC地址"
COL_MAINTENANCE_SET_POINT_COMM = "维护值设定点位_通讯地址"
COL_MAINTENANCE_ENABLE_POINT = "维护使能开关点位"
COL_MAINTENANCE_ENABLE_POINT_PLC = "维护使能开关点位_PLC地址"
COL_MAINTENANCE_ENABLE_POINT_COMM = "维护使能开关点位_通讯地址"

def _is_cell_empty(cell_value) -> bool:
    """检查单元格值是否为空"""
    return cell_value is None or (isinstance(cell_value, str) and not cell_value.strip())

def _clear_cell_value(sheet, row_idx: int, col_idx: int):
    """清空指定单元格的值"""
    if col_idx:  # 确保列索引存在
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.value = None

def generate_fat_checklist_from_source(original_file_path: str, output_dir: str, output_filename: str) -> Tuple[bool, Optional[str], Optional[str]]:
    """
    通过读取原始上传的IO点表文件，处理预留点位和设定值逻辑后，生成FAT点检表。
    此方法使用 openpyxl 以尽量保留原始格式和公式。
    
    处理逻辑：
    1. 预留点位处理：如果"变量名称（HMI）"和"变量描述"为空，则根据"通道位号"自动填充它们。
    2. 设定值处理：
       - SLL设定值为空时，删除SLL设定点位、SLL设定点位_PLC地址、SLL设定点位_通讯地址、LL报警
       - SL设定值为空时，删除SL设定点位、SL设定点位_PLC地址、SL设定点位_通讯地址、L报警
       - SH设定值为空时，删除SH设定点位、SH设定点位_PLC地址、SH设定点位_通讯地址、H报警
       - SHH设定值为空时，删除SHH设定点位、SHH设定点位_PLC地址、SHH设定点位_通讯地址、HH报警
    3. 维护点位处理：
       - 如果SLL、SL、SH、SHH设定值任意一个有数据，则保留维护相关点位
       - 如果SLL、SL、SH、SHH设定值全为空，则删除维护值设定点位、维护值设定点位_PLC地址、
         维护值设定点位_通讯地址、维护使能开关点位、维护使能开关点位_PLC地址、维护使能开关点位_通讯地址

    Args:
        original_file_path (str): 原始已验证IO点表Excel文件的路径 (应为 .xlsx)。
        output_dir (str): FAT点检表应保存的目录。
        output_filename (str): 生成的FAT点检表的输出文件名 (应为 .xlsx)。

    Returns:
        Tuple[bool, Optional[str], Optional[str]]: (成功状态, 生成文件的路径, 错误消息)
    """
    if not original_file_path or not os.path.exists(original_file_path):
        error_msg = f"原始IO点表文件路径 '{original_file_path}' 无效或文件不存在。"
        logger.error(f"FAT生成失败 (原始文件问题): {error_msg}")
        return False, None, error_msg
    
    if not original_file_path.endswith('.xlsx'):
        error_msg = f"仅支持 .xlsx 格式的Excel文件以保留格式和公式。文件: {original_file_path}"
        logger.error(error_msg)
        return False, None, error_msg

    if not output_dir:
        error_msg = "未提供FAT点检表的输出目录。"
        logger.error(f"FAT生成失败 (输出目录问题): {error_msg}")
        return False, None, error_msg
        
    if not output_filename:
        error_msg = "未提供FAT点检表的输出文件名。"
        logger.error(f"FAT生成失败 (输出文件名问题): {error_msg}")
        return False, None, error_msg

    destination_path = os.path.join(output_dir, output_filename)

    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(original_file_path)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            logger.info(f"正在处理工作表: {sheet_name} 进行FAT点表预留点位处理 (使用openpyxl)")

            # 寻找标题行并映射列名到列索引 (1-based)
            header_row = None
            col_map = {}
            # 通常标题在第一行，但为了稳健可以搜索几行
            for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(5, sheet.max_row), values_only=False)):
                temp_map = {}
                possible_header = False
                for c_idx, cell in enumerate(row):
                    if cell.value:
                        # 基础列
                        if cell.value == COL_HMI_VARIABLE_NAME: temp_map[COL_HMI_VARIABLE_NAME] = c_idx + 1; possible_header=True
                        if cell.value == COL_VARIABLE_DESCRIPTION: temp_map[COL_VARIABLE_DESCRIPTION] = c_idx + 1; possible_header=True
                        if cell.value == COL_CHANNEL_TAG: temp_map[COL_CHANNEL_TAG] = c_idx + 1; possible_header=True
                        
                        # 设定值列
                        if cell.value == COL_SLL_SET_VALUE: temp_map[COL_SLL_SET_VALUE] = c_idx + 1
                        if cell.value == COL_SL_SET_VALUE: temp_map[COL_SL_SET_VALUE] = c_idx + 1
                        if cell.value == COL_SH_SET_VALUE: temp_map[COL_SH_SET_VALUE] = c_idx + 1
                        if cell.value == COL_SHH_SET_VALUE: temp_map[COL_SHH_SET_VALUE] = c_idx + 1
                        
                        # 设定点位及地址列
                        if cell.value == COL_SLL_SET_POINT: temp_map[COL_SLL_SET_POINT] = c_idx + 1
                        if cell.value == COL_SLL_SET_POINT_PLC: temp_map[COL_SLL_SET_POINT_PLC] = c_idx + 1
                        if cell.value == COL_SLL_SET_POINT_COMM: temp_map[COL_SLL_SET_POINT_COMM] = c_idx + 1
                        if cell.value == COL_SL_SET_POINT: temp_map[COL_SL_SET_POINT] = c_idx + 1
                        if cell.value == COL_SL_SET_POINT_PLC: temp_map[COL_SL_SET_POINT_PLC] = c_idx + 1
                        if cell.value == COL_SL_SET_POINT_COMM: temp_map[COL_SL_SET_POINT_COMM] = c_idx + 1
                        if cell.value == COL_SH_SET_POINT: temp_map[COL_SH_SET_POINT] = c_idx + 1
                        if cell.value == COL_SH_SET_POINT_PLC: temp_map[COL_SH_SET_POINT_PLC] = c_idx + 1
                        if cell.value == COL_SH_SET_POINT_COMM: temp_map[COL_SH_SET_POINT_COMM] = c_idx + 1
                        if cell.value == COL_SHH_SET_POINT: temp_map[COL_SHH_SET_POINT] = c_idx + 1
                        if cell.value == COL_SHH_SET_POINT_PLC: temp_map[COL_SHH_SET_POINT_PLC] = c_idx + 1
                        if cell.value == COL_SHH_SET_POINT_COMM: temp_map[COL_SHH_SET_POINT_COMM] = c_idx + 1
                        
                        # 报警列
                        if cell.value == COL_LL_ALARM: temp_map[COL_LL_ALARM] = c_idx + 1
                        if cell.value == COL_LL_ALARM_PLC: temp_map[COL_LL_ALARM_PLC] = c_idx + 1
                        if cell.value == COL_LL_ALARM_COMM: temp_map[COL_LL_ALARM_COMM] = c_idx + 1
                        if cell.value == COL_L_ALARM: temp_map[COL_L_ALARM] = c_idx + 1
                        if cell.value == COL_L_ALARM_PLC: temp_map[COL_L_ALARM_PLC] = c_idx + 1
                        if cell.value == COL_L_ALARM_COMM: temp_map[COL_L_ALARM_COMM] = c_idx + 1
                        if cell.value == COL_H_ALARM: temp_map[COL_H_ALARM] = c_idx + 1
                        if cell.value == COL_H_ALARM_PLC: temp_map[COL_H_ALARM_PLC] = c_idx + 1
                        if cell.value == COL_H_ALARM_COMM: temp_map[COL_H_ALARM_COMM] = c_idx + 1
                        if cell.value == COL_HH_ALARM: temp_map[COL_HH_ALARM] = c_idx + 1
                        if cell.value == COL_HH_ALARM_PLC: temp_map[COL_HH_ALARM_PLC] = c_idx + 1
                        if cell.value == COL_HH_ALARM_COMM: temp_map[COL_HH_ALARM_COMM] = c_idx + 1
                        
                        # 维护相关列
                        if cell.value == COL_MAINTENANCE_SET_POINT: temp_map[COL_MAINTENANCE_SET_POINT] = c_idx + 1
                        if cell.value == COL_MAINTENANCE_SET_POINT_PLC: temp_map[COL_MAINTENANCE_SET_POINT_PLC] = c_idx + 1
                        if cell.value == COL_MAINTENANCE_SET_POINT_COMM: temp_map[COL_MAINTENANCE_SET_POINT_COMM] = c_idx + 1
                        if cell.value == COL_MAINTENANCE_ENABLE_POINT: temp_map[COL_MAINTENANCE_ENABLE_POINT] = c_idx + 1
                        if cell.value == COL_MAINTENANCE_ENABLE_POINT_PLC: temp_map[COL_MAINTENANCE_ENABLE_POINT_PLC] = c_idx + 1
                        if cell.value == COL_MAINTENANCE_ENABLE_POINT_COMM: temp_map[COL_MAINTENANCE_ENABLE_POINT_COMM] = c_idx + 1
                        
                if COL_HMI_VARIABLE_NAME in temp_map and COL_VARIABLE_DESCRIPTION in temp_map and COL_CHANNEL_TAG in temp_map:
                    col_map = temp_map
                    header_row = r_idx + 1
                    logger.info(f"在工作表 '{sheet_name}' 的第 {header_row} 行找到标题行。基础列映射: HMI变量名, 变量描述, 通道位号")
                    break
            
            if not header_row or not col_map:
                logger.warning(f"工作表 '{sheet_name}' 未能找到包含所有必需列的标题行: {COL_HMI_VARIABLE_NAME}, {COL_VARIABLE_DESCRIPTION}, {COL_CHANNEL_TAG}. 该工作表将不被处理预留点位。")
                continue

            modified_count = 0
            set_value_processed_count = 0
            maintenance_processed_count = 0
            
            # 从标题行之后开始遍历数据行
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                hmi_cell = sheet.cell(row=row_idx, column=col_map[COL_HMI_VARIABLE_NAME])
                desc_cell = sheet.cell(row=row_idx, column=col_map[COL_VARIABLE_DESCRIPTION])
                channel_cell = sheet.cell(row=row_idx, column=col_map[COL_CHANNEL_TAG])

                hmi_val = hmi_cell.value
                desc_val = desc_cell.value
                channel_val = channel_cell.value

                is_hmi_empty = hmi_val is None or (isinstance(hmi_val, str) and not hmi_val.strip())
                is_desc_empty = desc_val is None or (isinstance(desc_val, str) and not desc_val.strip())
                is_channel_valid = channel_val is not None and str(channel_val).strip() != ""

                # 处理预留点位（原有逻辑）
                if is_hmi_empty and is_desc_empty and is_channel_valid:
                    channel_tag_str = str(channel_val).strip()
                    hmi_cell.value = f"YLDW{channel_tag_str}"
                    desc_cell.value = f"{channel_tag_str}预留点位"
                    modified_count += 1

                # 处理设定值相关逻辑
                # 获取所有设定值（如果列不存在则视为空）
                sll_set_value = None
                if COL_SLL_SET_VALUE in col_map:
                    sll_set_value = sheet.cell(row=row_idx, column=col_map[COL_SLL_SET_VALUE]).value
                
                sl_set_value = None
                if COL_SL_SET_VALUE in col_map:
                    sl_set_value = sheet.cell(row=row_idx, column=col_map[COL_SL_SET_VALUE]).value
                
                sh_set_value = None
                if COL_SH_SET_VALUE in col_map:
                    sh_set_value = sheet.cell(row=row_idx, column=col_map[COL_SH_SET_VALUE]).value
                
                shh_set_value = None
                if COL_SHH_SET_VALUE in col_map:
                    shh_set_value = sheet.cell(row=row_idx, column=col_map[COL_SHH_SET_VALUE]).value

                # 检查SLL设定值，如果为空则删除相关字段
                if _is_cell_empty(sll_set_value):
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SLL_SET_POINT))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SLL_SET_POINT_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SLL_SET_POINT_COMM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_LL_ALARM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_LL_ALARM_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_LL_ALARM_COMM))

                # 检查SL设定值，如果为空则删除相关字段
                if _is_cell_empty(sl_set_value):
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SL_SET_POINT))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SL_SET_POINT_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SL_SET_POINT_COMM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_L_ALARM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_L_ALARM_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_L_ALARM_COMM))

                # 检查SH设定值，如果为空则删除相关字段
                if _is_cell_empty(sh_set_value):
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SH_SET_POINT))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SH_SET_POINT_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SH_SET_POINT_COMM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_H_ALARM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_H_ALARM_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_H_ALARM_COMM))

                # 检查SHH设定值，如果为空则删除相关字段
                if _is_cell_empty(shh_set_value):
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SHH_SET_POINT))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SHH_SET_POINT_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_SHH_SET_POINT_COMM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_HH_ALARM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_HH_ALARM_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_HH_ALARM_COMM))
                
                # 统计处理的行数
                if (_is_cell_empty(sll_set_value) or _is_cell_empty(sl_set_value) or 
                    _is_cell_empty(sh_set_value) or _is_cell_empty(shh_set_value)):
                    set_value_processed_count += 1

                # 处理维护相关点位逻辑
                # 检查是否所有设定值都为空
                all_set_values_empty = (
                    _is_cell_empty(sll_set_value) and 
                    _is_cell_empty(sl_set_value) and 
                    _is_cell_empty(sh_set_value) and 
                    _is_cell_empty(shh_set_value)
                )
                
                if all_set_values_empty:
                    # 如果所有设定值都为空，删除维护相关字段
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_MAINTENANCE_SET_POINT))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_MAINTENANCE_SET_POINT_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_MAINTENANCE_SET_POINT_COMM))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_MAINTENANCE_ENABLE_POINT))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_MAINTENANCE_ENABLE_POINT_PLC))
                    _clear_cell_value(sheet, row_idx, col_map.get(COL_MAINTENANCE_ENABLE_POINT_COMM))
                    maintenance_processed_count += 1
            
            # 输出处理统计信息
            if modified_count > 0:
                logger.info(f"在工作表 '{sheet_name}' 中处理了 {modified_count} 个预留点位。")
            else:
                logger.info(f"在工作表 '{sheet_name}' 中未找到或未处理预留点位。")
            
            if set_value_processed_count > 0:
                logger.info(f"在工作表 '{sheet_name}' 中处理了 {set_value_processed_count} 行的设定值相关字段清理。")
            
            if maintenance_processed_count > 0:
                logger.info(f"在工作表 '{sheet_name}' 中处理了 {maintenance_processed_count} 行的维护相关字段清理。")

        # 保存修改后的工作簿
        workbook.save(destination_path)
        
        logger.info(f"带预留点位处理的FAT点检表 (保留格式) 已成功生成于: {destination_path}")
        return True, destination_path, None

    except FileNotFoundError:
        error_msg = f"原始文件 '{original_file_path}' 未找到。"
        logger.error(error_msg)
        return False, None, error_msg
    except ImportError:
        error_msg = "处理Excel文件需要 openpyxl 库。请确保它已安装 (pip install openpyxl)。"
        logger.error(error_msg)
        return False, None, error_msg
    except Exception as e:
        error_msg = f"使用openpyxl处理Excel文件并生成FAT点检表时发生错误: {str(e)}"
        logger.error(error_msg, exc_info=True)
        return False, None, error_msg 