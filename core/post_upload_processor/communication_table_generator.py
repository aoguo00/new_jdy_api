import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from .uploaded_file_processor.io_data_model import UploadedIOPoint


def generate_communication_table_excel(output_path: str, io_points: List[UploadedIOPoint]) -> bool:
    """
    生成上下位通讯点表Excel文件。
    包含所有类型的点位：IO通道点位、第三方设备点位和中间点位。
    :param output_path: 输出文件的完整路径
    :param io_points: 从所有工作表解析出的 UploadedIOPoint 对象列表（包含所有类型点位）
    :return: 是否生成成功
    """
    # 表头字段，按截图顺序
    headers: List[str] = [
        "序号", "过程控制", "检测点名称", "信号范围", "数据范围", "单位", "信号类型", "数据类型", "供电", "PLC绝对地址", "上位机通讯地址", "低低报", "低报", "高报", "高高报", "备注"
    ]
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "上下位通讯点表"

        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 定义对齐样式（左对齐）
        left_alignment = Alignment(horizontal='left', vertical='center')

        # 添加表头
        ws.append(headers)

        # 设置表头样式
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.border = thin_border
            cell.alignment = left_alignment

        # 包含所有类型的点位：主IO点位、中间点位、第三方设备点位
        # 过滤掉无效的点位（没有HMI变量名的点位）
        all_valid_points = [
            p for p in io_points
            if p and p.hmi_variable_name and p.hmi_variable_name.strip()
        ]

        # 填充数据
        for index, point in enumerate(all_valid_points):
            serial_number = index + 1
            process_control_value = point.hmi_variable_name if point.hmi_variable_name else ""
            detection_point_name = point.variable_description if point.variable_description else ""

            signal_range_value = ""
            data_range_value = ""

            # 信号范围：只有硬点的AI/AO模块才显示"4~20mA"，软点为空
            if point.source_type == "main_io" and point.module_type in ["AI", "AO"]:
                signal_range_value = "4~20mA"

            # 数据范围：只有AI/AO类型的点位才处理量程范围
            if point.module_type in ["AI", "AO"]:
                # Handle Data Range for AI/AO
                low_limit = point.range_low_limit
                high_limit = point.range_high_limit

                if low_limit is not None and str(low_limit).strip() != "" and \
                   high_limit is not None and str(high_limit).strip() != "":
                    data_range_value = f"{str(low_limit).strip()}~{str(high_limit).strip()}"
                elif low_limit is not None and str(low_limit).strip() != "":
                    data_range_value = str(low_limit).strip()
                elif high_limit is not None and str(high_limit).strip() != "":
                    data_range_value = str(high_limit).strip()

            # 获取单位信息，确保处理None和空字符串的情况
            unit_value = ""
            if point.unit is not None:
                unit_str = str(point.unit).strip()
                if unit_str:
                    unit_value = unit_str

            # 确定信号类型：硬点显示模块类型，软点为空
            signal_type_value = ""
            if point.source_type == "main_io" and point.module_type:
                # 主IO硬件点位：显示模块类型（AI、AO、DI、DO）
                signal_type_value = point.module_type
            elif point.source_type in ["intermediate_from_main", "third_party"]:
                # 中间点位和第三方设备点位：为空
                signal_type_value = ""
            else:
                # 其他情况：显示模块类型（如果有的话）
                signal_type_value = point.module_type if point.module_type else ""

            # 获取数据类型
            data_type_value = ""
            if point.data_type is not None:
                data_type_str = str(point.data_type).strip()
                if data_type_str:
                    data_type_value = data_type_str

            # 获取PLC绝对地址和上位机通讯地址
            plc_address_value = ""
            if point.plc_absolute_address is not None:
                plc_addr_str = str(point.plc_absolute_address).strip()
                if plc_addr_str:
                    plc_address_value = plc_addr_str

            hmi_comm_address_value = ""
            if point.hmi_communication_address is not None:
                hmi_comm_addr_str = str(point.hmi_communication_address).strip()
                if hmi_comm_addr_str:
                    hmi_comm_address_value = hmi_comm_addr_str

            # 获取SLL、SL、SH、SHH设定值
            sll_set_value = ""
            if point.sll_set_value is not None:
                sll_str = str(point.sll_set_value).strip()
                if sll_str:
                    sll_set_value = sll_str

            sl_set_value = ""
            if point.sl_set_value is not None:
                sl_str = str(point.sl_set_value).strip()
                if sl_str:
                    sl_set_value = sl_str

            sh_set_value = ""
            if point.sh_set_value is not None:
                sh_str = str(point.sh_set_value).strip()
                if sh_str:
                    sh_set_value = sh_str

            shh_set_value = ""
            if point.shh_set_value is not None:
                shh_str = str(point.shh_set_value).strip()
                if shh_str:
                    shh_set_value = shh_str

            row_data = [
                serial_number,
                process_control_value,
                detection_point_name,
                signal_range_value,  # 信号范围
                data_range_value,    # 数据范围
                unit_value,  # 单位 - 从上传文件中的单位列获取
                signal_type_value,  # 信号类型 - 硬点显示模块类型，软点为空
                data_type_value,  # 数据类型 - 从上传文件中的数据类型列获取
                point.power_supply_type if point and point.power_supply_type else "",  # 供电
                plc_address_value,  # PLC绝对地址
                hmi_comm_address_value,  # 上位机通讯地址
                sll_set_value,  # 低低报
                sl_set_value,   # 低报
                sh_set_value,   # 高报
                shh_set_value,  # 高高报
                ""   # 备注 - 移到最后
            ]
            ws.append(row_data)

            # 设置当前行的样式（边框和左对齐）
            current_row = index + 2  # +2 因为表头占第1行，数据从第2行开始
            for col_num in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = thin_border
                cell.alignment = left_alignment

        # 设置列宽 - 为重要列设置固定宽度，其他列自适应
        column_widths = {
            1: 8,   # 序号
            2: 25,  # 过程控制 - 设置较宽以显示完整的HMI变量名
            3: 30,  # 检测点名称 - 设置较宽以显示完整的描述
            4: 12,  # 信号范围
            5: 15,  # 数据范围
            6: 10,  # 单位
            7: 12,  # 信号类型
            8: 12,  # 数据类型
            9: 10,  # 供电
            10: 18, # PLC绝对地址
            11: 18, # 上位机通讯地址
            12: 12, # 低低报
            13: 12, # 低报
            14: 12, # 高报
            15: 12, # 高高报
            16: 15  # 备注 - 移到最后
        }

        # 应用列宽设置
        for col_num, width in column_widths.items():
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

        wb.save(output_path)
        return True
    except Exception as e:
        print(f"生成上下位通讯点表失败: {e}")
        return False

# 中文注释：
# 1. 该函数会在指定路径生成一个Excel文件，sheet名为"上下位通讯点表"。
# 2. 包含所有类型的点位：IO通道点位、第三方设备点位和中间点位。
# 3. 所有单元格都设置了边框和左对齐格式。
# 4. 为每列设置了合适的列宽，确保内容完整显示（过程控制列25字符宽，检测点名称列30字符宽）。
# 5. 第一列为自动递增的序号。
# 6. 第二列 "过程控制" 来自传入的 io_points 列表中的 hmi_variable_name 属性。
# 7. 第三列 "检测点名称" 来自 variable_description 属性。
# 8. 第四列 "信号范围" 只对硬点的AI/AO模块自动填写为"4~20mA"，软点为空。
# 9. 第五列 "数据范围" 根据 range_low_limit 和 range_high_limit 自动生成。
# 10. 第六列 "单位" 来自上传文件中的 unit 字段。
# 11. 第七列 "信号类型" 根据点位类型确定：硬点显示模块类型（AI/AO/DI/DO），软点为空。
# 12. 第八列 "数据类型" 来自上传文件中的 data_type 字段。
# 13. 第九列 "供电" 来自 power_supply_type 属性。
# 14. 第十列 "PLC绝对地址" 来自 plc_absolute_address 属性。
# 15. 第十一列 "上位机通讯地址" 来自 hmi_communication_address 属性。
# 16. 第十二列 "低低报" 来自 sll_set_value 属性。
# 17. 第十三列 "低报" 来自 sl_set_value 属性。
# 18. 第十四列 "高报" 来自 sh_set_value 属性。
# 19. 第十五列 "高高报" 来自 shh_set_value 属性。
# 20. 第十六列 "备注" 当前为空，移到最后。