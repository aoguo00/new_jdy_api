# core/post_upload_processor/uploaded_file_processor/excel_reader.py
import openpyxl
import pandas as pd
from typing import List, Dict, Any, Optional, Tuple
from .io_data_model import UploadedIOPoint
import logging

logger = logging.getLogger(__name__)

# 主IO点表在Excel中通常的名称
MAIN_IO_SHEET_NAME = "IO点表"

# 这个映射表是核心，它将Excel中的中文表头映射到UploadedIOPoint数据类的属性名
# 我们需要确保这个映射是完整和正确的，基于 excel_exporter.py 中的 headers_plc
HEADER_TO_ATTRIBUTE_MAP: Dict[str, str] = {
    "序号": "serial_number",
    "模块名称": "module_name",
    "模块类型": "module_type",
    "供电类型（有源/无源）": "power_supply_type",
    "线制": "wiring_system",
    "通道位号": "channel_tag",
    "场站名": "site_name",
    "场站编号": "site_number",
    "变量名称（HMI）": "hmi_variable_name",
    "变量描述": "variable_description",
    "数据类型": "data_type",
    "单位": "unit",
    "保存历史": "save_history",
    "掉电保护": "power_off_protection",
    "量程低限": "range_low_limit",
    "量程高限": "range_high_limit",
    "SLL设定值": "sll_set_value",
    "SLL设定点位": "sll_set_point",
    "SLL设定点位_PLC地址": "sll_set_point_plc_address",
    "SLL设定点位_通讯地址": "sll_set_point_comm_address",
    "SL设定值": "sl_set_value",
    "SL设定点位": "sl_set_point",
    "SL设定点位_PLC地址": "sl_set_point_plc_address",
    "SL设定点位_通讯地址": "sl_set_point_comm_address",
    "SH设定值": "sh_set_value",
    "SH设定点位": "sh_set_point",
    "SH设定点位_PLC地址": "sh_set_point_plc_address",
    "SH设定点位_通讯地址": "sh_set_point_comm_address",
    "SHH设定值": "shh_set_value",
    "SHH设定点位": "shh_set_point",
    "SHH设定点位_PLC地址": "shh_set_point_plc_address",
    "SHH设定点位_通讯地址": "shh_set_point_comm_address",
    "LL报警": "ll_alarm",
    "LL报警_PLC地址": "ll_alarm_plc_address",
    "LL报警_通讯地址": "ll_alarm_comm_address",
    "L报警": "l_alarm",
    "L报警_PLC地址": "l_alarm_plc_address",
    "L报警_通讯地址": "l_alarm_comm_address",
    "H报警": "h_alarm",
    "H报警_PLC地址": "h_alarm_plc_address",
    "H报警_通讯地址": "h_alarm_comm_address",
    "HH报警": "hh_alarm",
    "HH报警_PLC地址": "hh_alarm_plc_address",
    "HH报警_通讯地址": "hh_alarm_comm_address",
    "维护值设定": "maintenance_set_value",
    "维护值设定点位": "maintenance_set_point",
    "维护值设定点位_PLC地址": "maintenance_set_point_plc_address",
    "维护值设定点位_通讯地址": "maintenance_set_point_comm_address",
    "维护使能开关点位": "maintenance_enable_switch_point",
    "维护使能开关点位_PLC地址": "maintenance_enable_switch_point_plc_address",
    "维护使能开关点位_通讯地址": "maintenance_enable_switch_point_comm_address",
    "PLC绝对地址": "plc_absolute_address",
    "上位机通讯地址": "hmi_communication_address"
}

# 新增：中间点位派生配置
# required_attrs_for_creation: 一个包含属性名元组的列表。每个元组代表一组"或"条件。
#                 例如 [('attr1', 'attr2'), ('attr3',)] 表示 (attr1存在 或 attr2存在) 且 (attr3存在)
#                 这里简化为：只要元组中任一属性有值，就认为该组条件满足。
#                 对于中间点，通常我们关心PLC地址或通讯地址至少有一个。
INTERMEDIATE_POINT_DEFINITIONS = [
    {
        'point_type_name': 'SLL设定点位',
        'hmi_name_attr': 'sll_set_point',
        'plc_addr_attr': 'sll_set_point_plc_address',
        'comm_addr_attr': 'sll_set_point_comm_address',
        'data_type': 'REAL', # SLL, SL, SH, SHH设定点通常是REAL
        'desc_suffix': '_SLL设定',
        'hmi_generation_suffix': '_LoLoLimit',
        'required_attrs_for_creation': ('sll_set_point_plc_address', 'sll_set_point_comm_address') # PLC或通讯地址至少有一个
    },
    {
        'point_type_name': 'SL设定点位',
        'hmi_name_attr': 'sl_set_point',
        'plc_addr_attr': 'sl_set_point_plc_address',
        'comm_addr_attr': 'sl_set_point_comm_address',
        'data_type': 'REAL',
        'desc_suffix': '_SL设定',
        'hmi_generation_suffix': '_LoLimit',
        'required_attrs_for_creation': ('sl_set_point_plc_address', 'sl_set_point_comm_address')
    },
    {
        'point_type_name': 'SH设定点位',
        'hmi_name_attr': 'sh_set_point',
        'plc_addr_attr': 'sh_set_point_plc_address',
        'comm_addr_attr': 'sh_set_point_comm_address',
        'data_type': 'REAL',
        'desc_suffix': '_SH设定',
        'hmi_generation_suffix': '_HiLimit',
        'required_attrs_for_creation': ('sh_set_point_plc_address', 'sh_set_point_comm_address')
    },
    {
        'point_type_name': 'SHH设定点位',
        'hmi_name_attr': 'shh_set_point',
        'plc_addr_attr': 'shh_set_point_plc_address',
        'comm_addr_attr': 'shh_set_point_comm_address',
        'data_type': 'REAL',
        'desc_suffix': '_SHH设定',
        'hmi_generation_suffix': '_HiHiLimit',
        'required_attrs_for_creation': ('shh_set_point_plc_address', 'shh_set_point_comm_address')
    },
    {
        'point_type_name': 'LL报警点位',
        'hmi_name_attr': 'll_alarm',
        'plc_addr_attr': 'll_alarm_plc_address',
        'comm_addr_attr': 'll_alarm_comm_address',
        'data_type': 'BOOL', # 报警点通常是BOOL
        'desc_suffix': '_LL报警',
        'hmi_generation_suffix': '_LL',
        'required_attrs_for_creation': ('ll_alarm_plc_address', 'll_alarm_comm_address')
    },
    {
        'point_type_name': 'L报警点位',
        'hmi_name_attr': 'l_alarm',
        'plc_addr_attr': 'l_alarm_plc_address',
        'comm_addr_attr': 'l_alarm_comm_address',
        'data_type': 'BOOL',
        'desc_suffix': '_L报警',
        'hmi_generation_suffix': '_L',
        'required_attrs_for_creation': ('l_alarm_plc_address', 'l_alarm_comm_address')
    },
    {
        'point_type_name': 'H报警点位',
        'hmi_name_attr': 'h_alarm',
        'plc_addr_attr': 'h_alarm_plc_address',
        'comm_addr_attr': 'h_alarm_comm_address',
        'data_type': 'BOOL',
        'desc_suffix': '_H报警',
        'hmi_generation_suffix': '_H',
        'required_attrs_for_creation': ('h_alarm_plc_address', 'h_alarm_comm_address')
    },
    {
        'point_type_name': 'HH报警点位',
        'hmi_name_attr': 'hh_alarm',
        'plc_addr_attr': 'hh_alarm_plc_address',
        'comm_addr_attr': 'hh_alarm_comm_address',
        'data_type': 'BOOL',
        'desc_suffix': '_HH报警',
        'hmi_generation_suffix': '_HH',
        'required_attrs_for_creation': ('hh_alarm_plc_address', 'hh_alarm_comm_address')
    },
    {
        'point_type_name': '维护值设定点位',
        'hmi_name_attr': 'maintenance_set_point',
        'plc_addr_attr': 'maintenance_set_point_plc_address',
        'comm_addr_attr': 'maintenance_set_point_comm_address',
        'data_type': 'REAL', # 假设维护值设定是模拟量
        'desc_suffix': '_维护设定',
        'hmi_generation_suffix': '_whz',
        'required_attrs_for_creation': ('maintenance_set_point_plc_address', 'maintenance_set_point_comm_address')
    },
    {
        'point_type_name': '维护使能开关点位',
        'hmi_name_attr': 'maintenance_enable_switch_point',
        'plc_addr_attr': 'maintenance_enable_switch_point_plc_address',
        'comm_addr_attr': 'maintenance_enable_switch_point_comm_address',
        'data_type': 'BOOL', # 使能开关是数字量
        'desc_suffix': '_维护使能',
        'hmi_generation_suffix': '_whzzt',
        'required_attrs_for_creation': ('maintenance_enable_switch_point_plc_address', 'maintenance_enable_switch_point_comm_address')
    }
]

# 新增：第三方表列名到 UploadedIOPoint 属性的映射
THIRD_PARTY_HEADER_TO_ATTRIBUTE_MAP: Dict[str, str] = {
    "场站名": "site_name", # 假设第三方表可能有这些信息
    "场站编号": "site_number",
    "变量名称": "hmi_variable_name", # 这是你图片中的列
    "变量描述": "variable_description", # 这是你图片中的列
    "数据类型": "data_type", # 这是你图片中的列
    "PLC地址": "plc_absolute_address", # 这是你图片中的列
    "MODBUS地址": "hmi_communication_address", # 这是你图片中的列
    "SLL设定值": "sll_set_value", # 注意，这些是 *值*，不是点位信息
    "SL设定值": "sl_set_value",
    "SH设定值": "sh_set_value",
    "SHH设定值": "shh_set_value",
    # 可以根据需要添加其他第三方表可能有的列
}


def _is_value_empty(value: Optional[str]) -> bool:
    """辅助函数：检查字符串值是否为空或仅包含空格。"""
    return not (value and value.strip())

def _clean_str(value: Any) -> Optional[str]:
    """将值转换为剥离空格的字符串，如果值为None则返回None。"""
    if value is None:
        return None
    return str(value).strip()

def _parse_io_sheet_to_uploaded_points(sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[UploadedIOPoint]:
    """
    将单个符合IO点表结构的工作表 (openpyxl.worksheet) 解析为 UploadedIOPoint 对象列表。
    现在会包含主点位及其派生的中间点位，并记录其来源信息。
    """
    all_parsed_points: List[UploadedIOPoint] = []
    sheet_title = sheet.title # 获取工作表标题以备后用

    header_row_cells = sheet[1]
    header_row = [_clean_str(cell.value) for cell in header_row_cells if cell.value is not None]

    current_file_header_to_attr_map: Dict[str, str] = {}
    for excel_header, attr_name in HEADER_TO_ATTRIBUTE_MAP.items():
        if excel_header in header_row:
            current_file_header_to_attr_map[excel_header] = attr_name
        else:
            logger.warning(f"主IO工作表 '{sheet_title}' 中，期望的表头 '{excel_header}' 未找到。该列数据将无法解析。")

    for row_idx, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        raw_row_data: Dict[str, Any] = {}
        is_empty_row = True
        for col_idx, cell in enumerate(row_cells):
            if col_idx < len(header_row):
                header_name = header_row[col_idx]
                if header_name in current_file_header_to_attr_map:
                    attribute_name = current_file_header_to_attr_map[header_name]
                    raw_row_data[attribute_name] = cell.value
                    if cell.value is not None and str(cell.value).strip() != "":
                        is_empty_row = False
            else:
                if cell.value is not None and str(cell.value).strip() != "":
                    is_empty_row = False

        if is_empty_row:
            logger.debug(f"主IO表 '{sheet_title}' 第 {row_idx} 行为空，跳过。")
            continue

        main_point_data: Dict[str, Optional[str]] = {
            attr: _clean_str(raw_row_data.get(attr)) for attr in HEADER_TO_ATTRIBUTE_MAP.values()
        }
        for field_name in UploadedIOPoint.__annotations__.keys():
            if field_name not in main_point_data:
                main_point_data[field_name] = None

        # 添加来源信息
        main_point_data['source_sheet_name'] = sheet_title
        main_point_data['source_type'] = "main_io"

        try:
            main_point = UploadedIOPoint(**main_point_data) # type: ignore
            if _is_value_empty(main_point.hmi_variable_name):
                # 检查是否是预留点，即HMI名称为空，但通道位号和PLC绝对地址不为空
                if not _is_value_empty(main_point.channel_tag) and not _is_value_empty(main_point.plc_absolute_address):
                    # site_number_str = main_point.site_number if main_point.site_number else "" # 场站编号暂时不用于HMI命名
                    channel_tag_str = main_point.channel_tag # 通道位号应该是必须的

                    # 新的HMI名称生成规则："YLDW" + 通道位号 (已移除场站编号)
                    if channel_tag_str: # 确保通道位号存在才进行拼接
                        main_point.hmi_variable_name = f"YLDW{channel_tag_str}"
                        main_point.variable_description = f"预留点位_{channel_tag_str}" # 描述保持不变或按需调整
                        # logger.info(f"主IO表行 {row_idx}: HMI名称为空，视为预留点，按规则 'YLDW{channel_tag_str}' 生成名称: {main_point.hmi_variable_name}")
                    else:
                        # 如果通道位号缺失，记录警告
                        default_hmi_name_if_missing_parts = f"TEMP_RESERVED_Row{row_idx}" # 使用行号作为备用标识
                        main_point.hmi_variable_name = default_hmi_name_if_missing_parts
                        main_point.variable_description = f"预留点位_Row{row_idx}" # 描述也使用行号
                        # logger.warning(f"主IO表行 {row_idx}: HMI名称为空，且通道位号缺失，预留点HMI名称按临时规则生成为: {main_point.hmi_variable_name}")
                else:
                    logger.warning(f"主IO表行 {row_idx}: HMI名称为空，且通道位号或PLC地址也可能不足以构成预留点，点位可能无效。HMI: {main_point.hmi_variable_name}, PLC: {main_point.plc_absolute_address}, CH: {main_point.channel_tag}")

            if not (_is_value_empty(main_point.hmi_variable_name) and _is_value_empty(main_point.plc_absolute_address) and _is_value_empty(main_point.hmi_communication_address)):
                all_parsed_points.append(main_point)
            else:
                logger.debug(f"主IO表行 {row_idx}: 主点位因HMI名、PLC地址、通讯地址均为空而被跳过。")
                continue

            for definition in INTERMEDIATE_POINT_DEFINITIONS:
                ip_hmi_name_val = getattr(main_point, definition['hmi_name_attr'], None)
                ip_plc_addr_val = getattr(main_point, definition['plc_addr_attr'], None)
                ip_comm_addr_val = getattr(main_point, definition['comm_addr_attr'], None)

                is_valid_intermediate_point = False
                for req_attr_key_for_main in definition['required_attrs_for_creation']:
                    if not _is_value_empty(getattr(main_point, req_attr_key_for_main, None)):
                        is_valid_intermediate_point = True
                        break

                if is_valid_intermediate_point:
                    intermediate_point_dict = {
                        'serial_number': main_point.serial_number,
                        'module_name': main_point.module_name,
                        'module_type': main_point.module_type,
                        'site_name': main_point.site_name,
                        'site_number': main_point.site_number,
                        'channel_tag': main_point.channel_tag,
                        'data_type': definition['data_type'],
                        'plc_absolute_address': _clean_str(ip_plc_addr_val),
                        'hmi_communication_address': _clean_str(ip_comm_addr_val),
                        'source_sheet_name': sheet_title, # 中间点也源自此主表
                        'source_type': "intermediate_from_main" # 标记为派生自主要IO
                    }
                    for field_name in UploadedIOPoint.__annotations__.keys():
                        if field_name not in intermediate_point_dict:
                            intermediate_point_dict[field_name] = None

                    # --- HMI 名称处理逻辑 (使用新的 hmi_generation_suffix 字段) ---
                    generated_hmi_name = None
                    # 优先使用 main_point.hmi_variable_name，因为此时它已经被正确设置了（无论是来自Excel还是生成的预留点名称）
                    if not _is_value_empty(main_point.hmi_variable_name):
                        base_name = main_point.hmi_variable_name
                        hmi_suffix = definition.get('hmi_generation_suffix')

                        if hmi_suffix: # 确保hmi_generation_suffix在定义中存在
                            # 直接拼接，父点HMI名 + 后缀
                            generated_hmi_name = f"{base_name}{hmi_suffix}"
                        else:
                            # 如果 INTERMEDIATE_POINT_DEFINITIONS 中没有定义 hmi_generation_suffix (理论上不应发生，因为我们刚添加了)
                            logger.error(f"主IO表行 {row_idx}: 中间点类型 '{definition['point_type_name']}' 严重错误 - 缺少 'hmi_generation_suffix' 定义。将尝试使用desc_suffix备用。")
                            generated_hmi_name = f"{base_name}{definition['desc_suffix']}" # 使用描述后缀作为非常规备用

                    else:
                        # 此情况表示 main_point.hmi_variable_name 在主点处理后仍然为空。
                        # 这理论上不应该发生，因为我们已经为HMI名称为空的预留点生成了名称。
                        # 如果真的发生，说明主点命名逻辑有缺陷或数据异常。
                        default_error_name_prefix = "ERROR_HMI_MAIN_EMPTY"
                        generated_hmi_name = f"{default_error_name_prefix}_{definition.get('hmi_generation_suffix', definition['desc_suffix'])}_{main_point.channel_tag or f'Row{row_idx}'}"
                        logger.error(f"主IO表行 {row_idx}: 主点HMI名称为空或无效 ('{main_point.hmi_variable_name}')，导致中间点 '{definition['point_type_name']}' HMI名称生成异常: {generated_hmi_name}")

                    intermediate_point_dict['hmi_variable_name'] = generated_hmi_name
                    # --- HMI 名称处理逻辑结束 ---

                    # --- 变量描述处理逻辑 ---
                    # 变量描述仍然基于主点的描述 和 definition['desc_suffix']
                    desc_base = _clean_str(main_point.variable_description)

                    # 如果主IO点的描述为空，则尝试使用主IO点的HMI名称作为描述的基础
                    if _is_value_empty(desc_base) and not _is_value_empty(main_point.hmi_variable_name):
                        desc_base = main_point.hmi_variable_name

                    # 如果两者都为空（例如，一个完全空的预留点行，且主HMI名也因故未生成），则给一个通用描述基础
                    if _is_value_empty(desc_base):
                        # 使用主点位号或行号构建一个基础描述，避免完全为空
                        desc_base = f"通道_{main_point.channel_tag or f'Row{row_idx}'}"

                    current_desc_suffix = definition['desc_suffix']
                    # 避免重复添加描述后缀，例如主描述已经是 "XXX_SLL设定"
                    if desc_base and current_desc_suffix and desc_base.endswith(current_desc_suffix):
                        intermediate_point_dict['variable_description'] = desc_base
                    else:
                        intermediate_point_dict['variable_description'] = f"{desc_base}{current_desc_suffix}"
                    # --- 变量描述处理逻辑结束 ---

                    ip_obj = UploadedIOPoint(**intermediate_point_dict) # type: ignore
                    all_parsed_points.append(ip_obj)
                    logger.debug(f"主IO表行 {row_idx}: 从主点 '{main_point.hmi_variable_name}' 派生中间点 '{ip_obj.hmi_variable_name}' (PLC: {ip_obj.plc_absolute_address}, Comm: {ip_obj.hmi_communication_address}) Source: {ip_obj.source_sheet_name}/{ip_obj.source_type}")

        except TypeError as e:
            logger.error(f"在主IO工作表 '{sheet_title}' 创建 UploadedIOPoint 实例时出错 (行 {row_idx}): {e}. 数据: {main_point_data if 'main_point_data' in locals() else raw_row_data}")
        except Exception as ex:
            logger.error(f"在主IO工作表 '{sheet_title}' 处理行 {row_idx} 时发生未知错误: {ex}. 数据: {main_point_data if 'main_point_data' in locals() else raw_row_data}", exc_info=True)

    logger.info(f"成功从主IO工作表 '{sheet_title}' 读取并处理了 {len(all_parsed_points)} 条IO点数据 (包括派生的中间点)。")
    return all_parsed_points

def _parse_third_party_df_to_uploaded_points(df: pd.DataFrame, sheet_name: str) -> List[UploadedIOPoint]:
    """将单个第三方设备DataFrame解析为UploadedIOPoint对象列表，并记录其来源信息。"""
    processed_data: List[UploadedIOPoint] = []
    if df.empty:
        logger.info(f"第三方工作表 '{sheet_name}' 为空，不进行处理。")
        return processed_data

    for index, row in df.iterrows():
        point_data_dict: Dict[str, Optional[str]] = {}
        for field_name in UploadedIOPoint.__annotations__.keys():
            point_data_dict[field_name] = None

        has_data = False
        for df_header, attr_name in THIRD_PARTY_HEADER_TO_ATTRIBUTE_MAP.items():
            if df_header in df.columns:
                cell_value = row.get(df_header)
                cleaned_value = _clean_str(cell_value if not pd.isna(cell_value) else None)
                point_data_dict[attr_name] = cleaned_value
                if not _is_value_empty(cleaned_value):
                    has_data = True

        if not has_data:
            logger.debug(f"第三方表 '{sheet_name}' 行 {index + 2} 数据全为空或无效，跳过。")
            continue

        # 添加来源信息
        point_data_dict['source_sheet_name'] = sheet_name
        point_data_dict['source_type'] = "third_party"

        if _is_value_empty(point_data_dict.get('hmi_variable_name')) and \
           _is_value_empty(point_data_dict.get('plc_absolute_address')) and \
           _is_value_empty(point_data_dict.get('hmi_communication_address')):
            logger.warning(f"第三方表 '{sheet_name}' 行 {index + 2}: HMI变量名、PLC地址和通讯地址均为空，跳过此点。 数据: {row.to_dict()}")
            continue

        try:
            point = UploadedIOPoint(**point_data_dict) # type: ignore
            processed_data.append(point)
            logger.debug(f"第三方表 '{sheet_name}' 行 {index + 2}: 解析点 '{point.hmi_variable_name}' Source: {point.source_sheet_name}/{point.source_type}")
        except TypeError as e:
            logger.error(f"在第三方工作表 '{sheet_name}' 创建 UploadedIOPoint 实例时出错 (行 {index + 2}): {e}. 数据: {point_data_dict}")
        except Exception as ex:
            logger.error(f"在第三方工作表 '{sheet_name}' 处理行 {index + 2} 时发生未知错误: {ex}. 数据: {point_data_dict}", exc_info=True)

    logger.info(f"成功从第三方工作表 '{sheet_name}' 读取并处理了 {len(processed_data)} 条IO点数据。")
    return processed_data


def load_workbook_data(file_path: str) -> Tuple[Dict[str, List[UploadedIOPoint]], Optional[str]]:
    """
    加载Excel工作簿中的所有数据。
    主IO点表 (默认为 "IO点表") 被解析，其点位 (包含派生中间点) 存入字典，键为该表名。
    其他所有工作表被视为第三方设备表，同样解析后，其点位列表以各自表名为键存入字典。
    最终返回一个按原始工作表名称组织的IO点数据字典。

    Args:
        file_path (str): Excel文件的路径。

    Returns:
        Tuple[Dict[str, List[UploadedIOPoint]], Optional[str]]:
            - points_by_sheet: 一个字典，键是原始工作表名，值是 UploadedIOPoint 对象列表。
            - error_message: 如果发生严重错误，则为错误消息字符串，否则为 None。
    """
    points_by_sheet: Dict[str, List[UploadedIOPoint]] = {}
    total_points_count = 0 # 用于日志记录总点位数

    try:
        opxl_workbook = openpyxl.load_workbook(file_path, data_only=True)

        try:
            pd_excel_file = pd.ExcelFile(file_path)
            all_sheet_names_from_pandas = pd_excel_file.sheet_names
        except Exception as e_pd:
            logger.error(f"Pandas无法打开Excel文件以获取工作表名称列表: {file_path}, 错误: {e_pd}")
            return {}, f"无法使用Pandas分析文件结构: {e_pd}"

        if MAIN_IO_SHEET_NAME in opxl_workbook.sheetnames:
            logger.info(f"找到主IO点表: '{MAIN_IO_SHEET_NAME}'。开始使用openpyxl解析 (包括中间点派生)...")
            io_sheet_obj = opxl_workbook[MAIN_IO_SHEET_NAME]
            main_and_intermediate_points = _parse_io_sheet_to_uploaded_points(io_sheet_obj)
            if main_and_intermediate_points: # 只有当列表非空时才添加
                points_by_sheet[MAIN_IO_SHEET_NAME] = main_and_intermediate_points
                total_points_count += len(main_and_intermediate_points)
                logger.info(f"主IO点表 '{MAIN_IO_SHEET_NAME}' 解析得到 {len(main_and_intermediate_points)} 个点位。")
            else:
                logger.info(f"主IO点表 '{MAIN_IO_SHEET_NAME}' 解析后数据为空 (包括派生点)。")
        else:
            logger.warning(f"在文件 '{file_path}' 中未找到预期主IO点表: '{MAIN_IO_SHEET_NAME}'。")

        for sheet_name in all_sheet_names_from_pandas:
            if sheet_name == MAIN_IO_SHEET_NAME:
                continue

            logger.info(f"尝试将工作表 '{sheet_name}' 作为第三方设备表加载 (使用pandas读取，然后转换为UploadedIOPoint)...")
            try:
                # 确保即使 df 为空，_parse_third_party_df_to_uploaded_points 也能安全处理并返回空列表
                df = pd.read_excel(pd_excel_file, sheet_name=sheet_name, header=0, dtype=str)
                third_party_points = _parse_third_party_df_to_uploaded_points(df, sheet_name)
                if third_party_points: # 只有当列表非空时才添加
                    points_by_sheet[sheet_name] = third_party_points
                    total_points_count += len(third_party_points)
                    logger.info(f"第三方工作表 '{sheet_name}' 解析得到 {len(third_party_points)} 个点位。")
                else:
                    # 检查是否真的为空，还是读取问题
                    df_check_actual_rows = pd.read_excel(pd_excel_file, sheet_name=sheet_name)
                    if df_check_actual_rows.empty:
                         logger.info(f"工作表 '{sheet_name}' (第三方) 为空或只有表头，已跳过。")
                    else:
                         logger.warning(f"工作表 '{sheet_name}' (第三方) 初始pandas读取(dtype=str)为空，但重读有内容。可能是非字符串数据导致。已作为空列表处理。")

            except Exception as e_read_tp:
                logger.warning(f"处理第三方工作表 '{sheet_name}' 时出错: {e_read_tp}。将跳过此表。", exc_info=True)

        logger.info(f"Excel文件 '{file_path}' 加载完成。总共从 {len(points_by_sheet)} 个工作表解析得到 {total_points_count} 个IO点对象。")
        return points_by_sheet, None

    except FileNotFoundError:
        logger.error(f"Excel文件未找到: {file_path}")
        return {}, f"Excel文件未找到: {file_path}"
    except Exception as e_global:
        logger.error(f"打开或处理Excel文件 '{file_path}' 时发生全局错误: {e_global}", exc_info=True)
        return {}, f"打开或处理Excel文件时发生错误: {e_global}"

if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s')

    test_file_path = "test_unified_io_table.xlsx"
    logger.info(f"准备创建统一化解析测试Excel文件: {test_file_path}")
    wb_test = openpyxl.Workbook()

    ws_io = wb_test.active
    if ws_io is not None:
        ws_io.title = MAIN_IO_SHEET_NAME
        io_headers = list(HEADER_TO_ATTRIBUTE_MAP.keys())
        ws_io.append(io_headers)

        ai_point_data = [""] * len(io_headers)
        def set_val(header_name, value):
            if header_name in io_headers:
                ai_point_data[io_headers.index(header_name)] = value

        set_val("序号", "1")
        set_val("模块名称", "AI_Module_01")
        set_val("模块类型", "AI")
        set_val("场站名", "测试主站")
        set_val("场站编号", "S001")
        set_val("变量名称（HMI）", "MAIN_AI_001")
        set_val("变量描述", "主AI点位001")
        set_val("数据类型", "REAL")
        set_val("PLC绝对地址", "%MD100")
        set_val("上位机通讯地址", "40100")
        set_val("通道位号", "CH01")
        set_val("SLL设定点位", "MAIN_AI_001_SLL_SP")
        set_val("SLL设定点位_PLC地址", "%MD102")
        set_val("SL设定点位_通讯地址", "40104")
        set_val("SH设定点位", "MAIN_AI_001_SH_SP")
        set_val("SH设定点位_PLC地址", "%MD106")
        set_val("SH设定点位_通讯地址", "40106")
        set_val("LL报警", "MAIN_AI_001_LL_ALM")
        set_val("LL报警_PLC地址", "%MX10.0")
        set_val("维护使能开关点位", "MAIN_AI_001_MAINT_EN")
        set_val("维护使能开关点位_通讯地址", "10010")
        ws_io.append(ai_point_data)

        di_point_data = [""] * len(io_headers)
        set_val("序号", "2"); set_val("模块类型", "DI"); set_val("变量名称（HMI）", "MAIN_DI_001");
        set_val("数据类型", "BOOL"); set_val("PLC绝对地址", "%MX12.0"); set_val("场站编号", "S001")
        ws_io.append(di_point_data)

        reserved_ai_data = [""] * len(io_headers)
        set_val("序号", "3"); set_val("模块类型", "AI"); set_val("PLC绝对地址", "%MD200");
        set_val("数据类型", "REAL"); set_val("通道位号", "CH02_RES");set_val("场站编号", "S001")
        set_val("SL设定点位_PLC地址", "%MD202")
        ws_io.append(reserved_ai_data)
        logger.info(f"'{MAIN_IO_SHEET_NAME}' 已准备数据。")

    ws_tp1 = wb_test.create_sheet(title="第三方设备A")
    tp_headers = ["变量名称", "变量描述", "数据类型", "PLC地址", "MODBUS地址", "SLL设定值", "场站编号"]
    ws_tp1.append(tp_headers)
    ws_tp1.append(["TP_BOOL_01", "第三方布尔1", "BOOL", "%MX500.0", "10500", None, "S001_TP"])
    ws_tp1.append(["TP_REAL_01", "第三方实数1", "REAL", "%MD600", "40600", "10.5", "S001_TP"])
    ws_tp1.append(["TP_NO_ADDR", "无地址第三方", "REAL", None, None, None, "S001_TP"])
    logger.info("'第三方设备A' sheet 已准备数据。")

    wb_test.create_sheet(title="空第三方")
    logger.info("'空第三方' sheet 已准备。")

    try:
        wb_test.save(test_file_path)
        logger.info(f"测试Excel文件 '{test_file_path}' 已成功创建/覆盖。")
    except Exception as e_save:
        logger.error(f"保存测试Excel文件时出错: {e_save}")

    logger.info(f"开始测试 load_workbook_data 函数，文件: {test_file_path}")
    loaded_data_dict, error_msg = load_workbook_data(test_file_path)

    if error_msg:
        logger.error(f"load_workbook_data 执行时返回错误: {error_msg}")
    else:
        logger.info(f"load_workbook_data 执行完毕。共解析 {len(loaded_data_dict)} 个工作表的数据。")
        total_loaded_points = 0
        for sheet_name, points_list in loaded_data_dict.items():
            logger.info(f"  工作表: '{sheet_name}' 包含 {len(points_list)} 个点位。")
            total_loaded_points += len(points_list)
            for i, point in enumerate(points_list):
                logger.debug(
                    f"    点 {i+1}: HMI='{point.hmi_variable_name}', "
                    f"Desc='{point.variable_description}', "
                    f"Type='{point.data_type}', "
                    f"PLC='{point.plc_absolute_address}', "
                    f"Comm='{point.hmi_communication_address}', "
                    f"SiteNo='{point.site_number}', "
                    f"Channel='{point.channel_tag}', "
                    f"SourceSheet='{point.source_sheet_name}', "
                    f"SourceType='{point.source_type}'"
                )
        logger.info(f"总共加载 {total_loaded_points} 个点位。")

        # 预期点位数量 (主IO表: 3主点 + 5(AI_001的中间点) + 1(预留AI的中间点) = 9点)
        # (第三方设备A: TP_BOOL_01, TP_REAL_01 = 2点; TP_NO_ADDR会被跳过)
        # (空第三方: 0点)
        # 总计 = 9 + 2 = 11
        expected_total_points = 11
        if total_loaded_points == expected_total_points:
            logger.info(f"总点位数量 ({total_loaded_points}) 符合预期 ({expected_total_points})。")
        else:
            logger.error(f"总点位数量 ({total_loaded_points}) 不符合预期 ({expected_total_points})！请检查日志。")

        expected_sheets_count = 3 # IO点表, 第三方设备A, 空第三方 (即使空也应该在keys里，value是空list)
        # 注意：如果_parse_third_party_df_to_uploaded_points在df为空时直接返回空列表，且我们只在列表非空时添加，
        # 那么 "空第三方" 可能不会出现在 loaded_data_dict 的键中。
        # 根据当前修改，如果df为空，third_party_points会是空列表，然后`if third_party_points:`会为false，所以"空第三方"不会被加入字典。
        # 因此，预期工作表数量应该是2个（IO点表，第三方设备A）。

        # 更新预期工作表数量，因为空表不会被加入字典
        expected_sheets_count_in_dict = 2
        if len(loaded_data_dict) == expected_sheets_count_in_dict:
            logger.info(f"解析的工作表数量 ({len(loaded_data_dict)}) 符合预期 ({expected_sheets_count_in_dict})。")
            if MAIN_IO_SHEET_NAME not in loaded_data_dict:
                 logger.error(f"主IO点表 '{MAIN_IO_SHEET_NAME}' 未在解析结果中找到！")
            if "第三方设备A" not in loaded_data_dict:
                 logger.error(f"第三方工作表 '第三方设备A' 未在解析结果中找到！")
            if "空第三方" in loaded_data_dict: # 这个不应该出现
                 logger.error(f"空工作表 '空第三方' 不应出现在解析结果的键中！")

        else:
            logger.error(f"解析的工作表数量 ({len(loaded_data_dict)}) 不符合预期 ({expected_sheets_count_in_dict})！Keys: {list(loaded_data_dict.keys())}")

